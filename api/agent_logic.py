import os
import json
import csv
import sqlite3
import logging
import re
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
from uuid import uuid4
from dotenv import load_dotenv
from pydantic import BaseModel, Field, field_validator
import traceback
import argparse
import asyncio

try:
    import pdfplumber
except ImportError:
    pdfplumber = None
    print("pdfplumber not installed. PDF text extraction will be disabled.")

try:
    from docx import Document
except ImportError:
    Document = None
    print("python-docx not installed. DOCX text extraction will be disabled.")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("openpyxl not installed. XLSX text extraction will be disabled.")


def _extract_text_from_pdf(file_path: str) -> str:
    if not pdfplumber:
        logger.warning("pdfplumber not available. Cannot extract text from PDF.")
        return ""
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
        return text
    except Exception as e:
        error_logger.error(f"Error extracting text from PDF {file_path}: {e}")
        return ""

def _extract_text_from_docx(file_path: str) -> str:
    if not Document:
        logger.warning("python-docx not available. Cannot extract text from DOCX.")
        return ""
    text = []
    try:
        document = Document(file_path)
        for para in document.paragraphs:
            text.append(para.text)
        return "\n".join(text)
    except Exception as e:
        error_logger.error(f"Error extracting text from DOCX {file_path}: {e}")
        return ""

def _extract_text_from_xlsx(file_path: str) -> str:
    if not openpyxl:
        logger.warning("openpyxl not available. Cannot extract text from XLSX.")
        return ""
    text = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text.append(str(cell.value))
        return "\n".join(text)
    except Exception as e:
        error_logger.error(f"Error extracting text from XLSX {file_path}: {e}")
        return ""

def _extract_text_from_csv(file_path: str) -> str:
    text = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                text.extend(row)
        return "\n".join(text)
    except Exception as e:
        error_logger.error(f"Error extracting text from CSV {file_path}: {e}")
        return ""

def _extract_text_from_json(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return json.dumps(data, indent=2)
    except Exception as e:
        error_logger.error(f"Error extracting text from JSON {file_path}: {e}")
        return ""

def _extract_text_from_txt(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        error_logger.error(f"Error extracting text from TXT {file_path}: {e}")
        return ""

from langchain_core.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain_core.output_parsers import StrOutputParser
from langchain_core.messages import HumanMessage, AIMessage
from langgraph.graph import StateGraph, END
from langchain_community.document_loaders import WebBaseLoader
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from supabase import create_client, Client as SupabaseClient
from langchain_google_genai import ChatGoogleGenerativeAI

import requests
from tenacity import retry, stop_after_attempt, wait_exponential
from cachetools import TTLCache

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    handlers=[logging.FileHandler("market_intelligence.log", mode="a"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

error_logger = logging.getLogger("agent_error_specific_logger")
error_file_handler = logging.FileHandler("market_intelligence_errors.log", mode="a")
error_file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s - %(message)s"))
error_logger.addHandler(error_file_handler)
error_logger.setLevel(logging.ERROR)
error_logger.propagate = False

if "USER_AGENT" not in os.environ:
    os.environ["USER_AGENT"] = "MarketIntelligenceAgent/1.0 (+http://example.com/botinfo)"
    logger.info(f"Default USER_AGENT set to: {os.environ['USER_AGENT']}")

load_dotenv()
search_results_cache = TTLCache(maxsize=100, ttl=3600)
_supabase_client_instance = None

def get_supabase_client() -> Optional[SupabaseClient]:
    global _supabase_client_instance
    if _supabase_client_instance:
        return _supabase_client_instance
    supabase_url = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    supabase_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_key:
        logger.warning("SUPABASE_SERVICE_ROLE_KEY not found, attempting fallback to SUPABASE_ANON_KEY.")
        supabase_key = os.getenv("SUPABASE_ANON_KEY")
    if supabase_url and supabase_key:
        try:
            url_domain_part = "URL_NOT_LOGGED"
            if "//" in supabase_url:
                url_domain_part = supabase_url.split("//")[-1].split(".")[0]
            logger.info(f"Initializing Supabase client for agent_logic with URL domain: {url_domain_part}...")
            
            _supabase_client_instance = create_client(supabase_url, supabase_key)
            logger.info("Supabase client for agent_logic initialized successfully.")
            return _supabase_client_instance
        except Exception as e:
            error_logger.error(f"Failed to initialize Supabase client: {e}\n{traceback.format_exc()}")
            _supabase_client_instance = None
            return None
    else:
        error_logger.error("CRITICAL: SUPABASE_URL or a SUPABASE_KEY not found.")
        _supabase_client_instance = None
        return None

def get_api_key(service_name: str, user_id: Optional[str] = None) -> Optional[str]:
    logger.debug(f"Attempting API key for service: {service_name}, UserID: {user_id or 'N/A'}")
    normalized_service_name = service_name.lower().replace("_", "").replace("api", "").replace("search", "").strip()
    if user_id:
        supabase = get_supabase_client()
        if supabase:
            try:
                response = supabase.table("data_sources").select("name, type, config, status").eq("user_id", user_id).eq("status", "active").execute()
                if response.data:
                    for source in response.data:
                        source_name_lower = (source.get("name") or "").lower().replace(" ", "")
                        source_type_lower = (source.get("type") or "").lower().replace(" ", "")
                        match_found = False
                        if normalized_service_name in source_name_lower:
                            match_found = True
                        elif normalized_service_name in source_type_lower:
                            match_found = True
                        elif (normalized_service_name in ["gemini", "googlegemini"] and
                              (any(term in source_type_lower for term in ["google", "gemini", "ai"]) or
                               any(term in source_name_lower for term in ["google", "gemini", "ai"]))):
                            match_found = True
                        if match_found:
                            config = source.get("config")
                            if config and isinstance(config, dict):
                                api_key_from_db = config.get("apiKey") or config.get("api_key")
                                if api_key_from_db and isinstance(api_key_from_db, str) and api_key_from_db.strip():
                                    logger.info(f"Using DB API key for: {service_name} (User: {user_id}, Source: {source.get('name')})")
                                    return api_key_from_db.strip()
                                else:
                                    logger.warning(f"DB source '{source.get('name')}' for {service_name} (User: {user_id}) missing/empty apiKey in config: {config}")
                            else:
                                logger.warning(f"DB source '{source.get('name')}' for {service_name} (User: {user_id}) config missing/invalid: {config}")
                logger.info(f"No user DB API key for: {service_name} (UserID: {user_id}). Fallback to env.")
            except Exception as e:
                error_logger.error(f"DB API key query error (service: {service_name}, user: {user_id}): {e}\n{traceback.format_exc()}")
        else:
            logger.warning("Supabase client NA for DB API key. Fallback to env.")
    internal_to_env_map = {
        "tavily": "TAVILY_API_KEY", "serpapi": "SERPAPI_API_KEY", "news_api": "NEWS_API_KEY", "newsapi": "NEWS_API_KEY",
        "financial_modeling_prep": "FINANCIAL_MODELING_PREP_API_KEY", "alpha_vantage": "ALPHA_VANTAGE_API_KEY",
        "mediastack": "MEDIASTACK_API_KEY", "google_gemini": os.getenv("GOOGLE_API_KEY_NAME", "GOOGLE_API_KEY"),
        "gemini": os.getenv("GOOGLE_API_KEY_NAME", "GOOGLE_API_KEY"), "google_genai": os.getenv("GOOGLE_API_KEY_NAME", "GOOGLE_API_KEY")}
    env_var_name = internal_to_env_map.get(service_name.upper()) or \
                   internal_to_env_map.get(normalized_service_name) or \
                   internal_to_env_map.get(service_name.lower())
    if not env_var_name:
        env_var_name = f"{service_name.upper()}_API_KEY"
    api_key_from_env = os.getenv(env_var_name)
    if api_key_from_env and api_key_from_env.strip():
        logger.info(f"Using env API key {env_var_name} for: {service_name}")
        return api_key_from_env.strip()
    if normalized_service_name in ["gemini", "googlegemini", "google_genai"] and not api_key_from_env:
        api_key_from_env = os.getenv("GOOGLE_API_KEY")
        if api_key_from_env and api_key_from_env.strip():
            logger.info(f"Using fallback GOOGLE_API_KEY for: {service_name}")
            return api_key_from_env.strip()
    logger.warning(f"API key not found for: {service_name} (UserID: {user_id or 'N/A'}, EnvVar: {env_var_name})")
    return None

def init_db():
    db_name = "market_intelligence_agent.db"
    db_path = ""
    try:
        if os.environ.get("VERCEL_ENV"):
            db_path = os.path.join("/tmp", db_name)
        else:
            api_python_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(api_python_dir, db_name)
        logger.info(f"Database path determined: {db_path}")
        conn = sqlite3.connect(db_path)
        cursor_obj = conn.cursor()
        cursor_obj.execute("CREATE TABLE IF NOT EXISTS states (id TEXT PRIMARY KEY, user_id TEXT, market_domain TEXT, query TEXT, state_data TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        cursor_obj.execute("CREATE INDEX IF NOT EXISTS idx_states_user_id ON states (user_id);")
        cursor_obj.execute("CREATE TABLE IF NOT EXISTS chat_history (session_id TEXT, message_type TEXT, content TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP, PRIMARY KEY (session_id, timestamp))")
        cursor_obj.execute("CREATE TABLE IF NOT EXISTS customer_insights (id TEXT PRIMARY KEY, state_id TEXT, segment_name TEXT, segment_data TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (state_id) REFERENCES states(id))")
        cursor_obj.execute("CREATE TABLE IF NOT EXISTS downloads (id TEXT PRIMARY KEY, state_id TEXT, category TEXT, file_path TEXT, file_type TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (state_id) REFERENCES states(id))")
        conn.commit()
        conn.close()
        logger.info(f"Database '{db_path}' initialized/verified successfully.")
    except Exception as e_db_init:
        error_logger.error(f"Failed to initialize database '{db_name}': {e_db_init} (Path: {db_path})")
        raise
init_db()

class MarketIntelligenceState(BaseModel):
    raw_news_data: List[Dict[str, Any]] = Field(default_factory=list)
    competitor_data: List[Dict[str, Any]] = Field(default_factory=list)
    financial_data: List[Dict] = Field(default_factory=list)
    market_trends: List[Dict[str, Any]] = Field(default_factory=list)
    opportunities: List[Dict[str, Any]] = Field(default_factory=list)
    strategic_recommendations: List[Dict[str, Any]] = Field(default_factory=list)
    customer_insights: List[Dict[str, Any]] = Field(default_factory=list)
    market_domain: str = "General Technology"
    query: Optional[str] = None
    user_id: Optional[str] = None
    question: Optional[str] = None
    query_response: Optional[str] = None
    report_template: Optional[str] = None
    vector_store_path: Optional[str] = None
    state_id: str = Field(default_factory=lambda: str(uuid4()))
    report_dir: Optional[str] = None
    chart_paths: List[str] = Field(default_factory=list)
    download_files: Dict[str, str] = Field(default_factory=dict)
    uploaded_files_content: List[Dict[str, Any]] = Field(default_factory=list)
    num_trends: Optional[int] = None
    num_opportunities: Optional[int] = None
    uploaded_document_ids: Optional[List[str]] = None # For passing IDs of pre-uploaded docs
    node_errors: List[Dict[str, str]] = Field(default_factory=list) # To track errors from individual nodes

    @field_validator("market_domain")
    @classmethod
    def validate_market_domain_value(cls, v_domain: str) -> str:
        if not v_domain:
            raise ValueError("Market domain cannot be empty.")
        if not re.match(r"^[a-zA-Z0-9\s-]+$", v_domain):
            raise ValueError("Market domain must contain only letters, numbers, spaces, or hyphens.")
        return v_domain.strip()

    @field_validator("query")
    @classmethod
    def validate_query_value(cls, v_query: Optional[str]) -> Optional[str]:
        if v_query and len(v_query.strip()) < 3:
            raise ValueError("Query must be at least 3 characters long if provided.")
        return v_query.strip() if v_query else None

    class Config:
        validate_assignment = True

def get_db_path():
    db_name = 'market_intelligence_agent.db'
    if os.environ.get("VERCEL_ENV"):
        return os.path.join("/tmp", db_name)
    else:
        api_python_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(api_python_dir, db_name)

def save_state(state_obj: MarketIntelligenceState):
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor_obj = conn.cursor()
        created_at_iso = datetime.now(timezone.utc).isoformat()
        logger.debug(f"Saving state for UserID: {state_obj.user_id}, StateID: {state_obj.state_id}, CreatedAt: {created_at_iso}")
        cursor_obj.execute('INSERT OR REPLACE INTO states (id, user_id, market_domain, query, state_data, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                           (state_obj.state_id, state_obj.user_id, state_obj.market_domain, state_obj.query, state_obj.model_dump_json(), created_at_iso))
        conn.commit()
        logger.info(f"State saved: ID={state_obj.state_id}, UserID={state_obj.user_id}, Domain='{state_obj.market_domain}' to {db_path}")
    except sqlite3.Error as e_save_sqlite:
        error_logger.error(f"SQLite error saving state {state_obj.state_id} for UserID {state_obj.user_id} to {db_path}: {e_save_sqlite}\n{traceback.format_exc()}")
    except Exception as e_save_state:
        error_logger.error(f"Unexpected error saving state {state_obj.state_id} for UserID {state_obj.user_id} to {db_path}: {e_save_state}\n{traceback.format_exc()}")
    finally:
        if conn:
            conn.close()

def list_user_analysis_states(user_id: str) -> List[Dict[str, Any]]:
    db_path = get_db_path()
    states_summary = []
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, market_domain, query, created_at, user_id FROM states WHERE user_id = ? ORDER BY DATETIME(created_at) DESC LIMIT 50",
            (user_id,)
        )
        rows = cursor.fetchall()

        column_names = [description[0] for description in cursor.description]

        for row in rows:
            row_dict = dict(zip(column_names, row))
            states_summary.append({
                "state_id": row_dict["id"],
                "market_domain": row_dict["market_domain"],
                "query": row_dict["query"],
                "created_at": row_dict["created_at"],
                "user_id": row_dict["user_id"]
            })
        logger.info(f"Fetched {len(states_summary)} analysis states for UserID {user_id}")
    except sqlite3.Error as e:
        error_logger.error(f"SQLite error fetching states for UserID {user_id}: {e}")
    except Exception as e:
        error_logger.error(f"Unexpected error fetching states for UserID {user_id}: {e}\n{traceback.format_exc()}")
    finally:
        if conn:
            conn.close()
    return states_summary

def get_state_download_info(state_id: str, user_id: str) -> Optional[Dict[str, Any]]:
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT query, market_domain, created_at, state_data, user_id FROM states WHERE id = ? AND user_id = ?",
            (state_id, user_id)
        )
        row = cursor.fetchone()

        if not row:
            logger.warning(f"Download Info: State ID {state_id} not found or not owned by user {user_id}")
            return None

        query, market_domain, created_at, state_data_json, db_user_id = row

        if db_user_id != user_id:
            logger.error(f"Download Info: Mismatch user_id for state {state_id}. Expected {user_id}, found {db_user_id}.")
            return None

        try:
            state_data_dict = json.loads(state_data_json)
        except json.JSONDecodeError:
            error_logger.error(f"Download Info: Failed to parse state_data for state {state_id}")
            return None

        downloadable_files_list = []

        agent_download_files = state_data_dict.get("download_files", {})
        if isinstance(agent_download_files, dict):
            for category, full_path in agent_download_files.items():
                if full_path and isinstance(full_path, str):
                     downloadable_files_list.append({
                        "category": category,
                        "filename": os.path.basename(full_path),
                        "description": f"{category.replace('_', ' ').title()} file"
                    })
                else:
                    logger.warning(f"Download Info: Invalid path for category '{category}' in state {state_id}: {full_path}")

        agent_chart_paths = state_data_dict.get("chart_paths", [])
        if isinstance(agent_chart_paths, list):
            for full_path in agent_chart_paths:
                if full_path and isinstance(full_path, str):
                    base = os.path.basename(full_path)
                    chart_name, _ = os.path.splitext(base)
                    category_name = f"chart_{chart_name.lower().replace(' ', '_').replace('-', '_')}"
                    downloadable_files_list.append({
                        "category": category_name,
                        "filename": base,
                        "description": f"Chart: {chart_name.replace('_', ' ').replace('-', ' ').title()}"
                    })
                else:
                     logger.warning(f"Download Info: Invalid chart path in state {state_id}: {full_path}")

        return {
            "state_id": state_id,
            "query": query,
            "market_domain": market_domain,
            "created_at": created_at,
            "files": downloadable_files_list
        }

    except sqlite3.Error as e:
        error_logger.error(f"Download Info: SQLite error for state {state_id}, user {user_id}: {e}")
        return None
    except Exception as e:
        error_logger.error(f"Download Info: Unexpected error for state {state_id}, user {user_id}: {e}\n{traceback.format_exc()}")
        return None
    finally:
        if conn:
            conn.close()

def get_download_file_path(state_id: str, user_id: str, file_identifier: str) -> Optional[str]:
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT state_data FROM states WHERE id = ? AND user_id = ?", # Only need state_data
            (state_id, user_id)
        )
        row = cursor.fetchone()
        if not row:
            logger.warning(f"File Download: State ID {state_id} not found or not owned by user {user_id}")
            return None

        state_data_json = row[0]
        try:
            state_data_dict = json.loads(state_data_json)
        except json.JSONDecodeError:
            error_logger.error(f"File Download: Failed to parse state_data for state {state_id}")
            return None

        target_path = None

        # Check in download_files (category lookup)
        agent_download_files = state_data_dict.get("download_files", {})
        if isinstance(agent_download_files, dict) and file_identifier in agent_download_files:
            target_path = agent_download_files[file_identifier]

        # If not found by category, check chart_paths (filename lookup)
        if not target_path:
            agent_chart_paths = state_data_dict.get("chart_paths", [])
            if isinstance(agent_chart_paths, list):
                for chart_path in agent_chart_paths:
                    if chart_path and isinstance(chart_path, str) and os.path.basename(chart_path) == file_identifier:
                        target_path = chart_path
                        break

        if not target_path or not isinstance(target_path, str):
            logger.warning(f"File Download: File identifier '{file_identifier}' not found in state {state_id} for user {user_id}.")
            return None

        # Security Check: Ensure the path is within the expected reports directory
        reports_base_dir = os.path.abspath(get_agent_base_reports_dir())
        if not isinstance(target_path, str): # Should be redundant given previous check, but good for safety
             error_logger.error(f"File Download: target_path is not a string for state {state_id}, identifier {file_identifier}.")
             return None
        resolved_target_path = os.path.abspath(target_path)

        if not resolved_target_path.startswith(reports_base_dir):
            error_logger.error(f"File Download SECURITY ALERT: Attempt to access path '{resolved_target_path}' outside base reports directory '{reports_base_dir}' for state {state_id}, user {user_id}.")
            return None

        if not os.path.exists(resolved_target_path) or not os.path.isfile(resolved_target_path):
            error_logger.error(f"File Download: File does not exist or is not a file at path '{resolved_target_path}' for state {state_id}, user {user_id}.")
            return None

        logger.info(f"File Download: Access validated for path '{resolved_target_path}' for state {state_id}, user {user_id}.")
        return resolved_target_path

    except sqlite3.Error as e:
        error_logger.error(f"File Download: SQLite error for state {state_id}, user {user_id}: {e}")
        return None
    except Exception as e:
        error_logger.error(f"File Download: Unexpected error for state {state_id}, user {user_id}: {e}\n{traceback.format_exc()}")
        return None
    finally:
        if conn:
            conn.close()

def list_user_analysis_states(user_id: str) -> List[Dict[str, Any]]:
    db_path = get_db_path()
    states_summary = []
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        # The DATETIME function is used to ensure correct sorting of ISO8601 strings.
        cursor.execute(
            "SELECT id, market_domain, query, created_at, user_id FROM states WHERE user_id = ? ORDER BY DATETIME(created_at) DESC LIMIT 50",
            (user_id,)
        )
        rows = cursor.fetchall()

        column_names = [description[0] for description in cursor.description]

        for row in rows:
            row_dict = dict(zip(column_names, row))
            states_summary.append({
                "state_id": row_dict["id"], # Map db 'id' to 'state_id' for consistency
                "market_domain": row_dict["market_domain"],
                "query": row_dict["query"],
                "created_at": row_dict["created_at"],
                "user_id": row_dict["user_id"]
            })
        logger.info(f"Fetched {len(states_summary)} analysis states for UserID {user_id}")
    except sqlite3.Error as e:
        error_logger.error(f"SQLite error fetching states for UserID {user_id}: {e}")
    except Exception as e:
        error_logger.error(f"Unexpected error fetching states for UserID {user_id}: {e}\n{traceback.format_exc()}")
    finally:
        if conn:
            conn.close()
    return states_summary

def get_state_download_info(state_id: str, user_id: str) -> Optional[Dict[str, Any]]:
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        # Ensure we select user_id to verify ownership, though it's used in WHERE too
        cursor.execute(
            "SELECT query, market_domain, created_at, state_data, user_id FROM states WHERE id = ? AND user_id = ?",
            (state_id, user_id)
        )
        row = cursor.fetchone()

        if not row:
            logger.warning(f"Download Info: State ID {state_id} not found or not owned by user {user_id}")
            return None

        query, market_domain, created_at, state_data_json, db_user_id = row

        if db_user_id != user_id:
            logger.error(f"Download Info: Mismatch user_id for state {state_id}. Expected {user_id}, found {db_user_id}.")
            return None

        try:
            state_data_dict = json.loads(state_data_json)
        except json.JSONDecodeError:
            error_logger.error(f"Download Info: Failed to parse state_data for state {state_id}")
            return None

        downloadable_files_list = []

        agent_download_files = state_data_dict.get("download_files", {})
        if isinstance(agent_download_files, dict):
            for category, full_path in agent_download_files.items():
                if full_path and isinstance(full_path, str):
                     downloadable_files_list.append({
                        "category": category,
                        "filename": os.path.basename(full_path),
                        "description": f"{category.replace('_', ' ').title()} file"
                    })
                else:
                    logger.warning(f"Download Info: Invalid path for category '{category}' in state {state_id}: {full_path}")

        agent_chart_paths = state_data_dict.get("chart_paths", [])
        if isinstance(agent_chart_paths, list):
            for full_path in agent_chart_paths:
                if full_path and isinstance(full_path, str):
                    base = os.path.basename(full_path)
                    chart_name, _ = os.path.splitext(base)
                    # Standardize chart category name
                    category_name = f"chart_{chart_name.lower().replace(' ', '_').replace('-', '_')}"
                    downloadable_files_list.append({
                        "category": category_name,
                        "filename": base,
                        "description": f"Chart: {chart_name.replace('_', ' ').replace('-', ' ').title()}"
                    })
                else:
                     logger.warning(f"Download Info: Invalid chart path in state {state_id}: {full_path}")

        return {
            "state_id": state_id,
            "query": query,
            "market_domain": market_domain,
            "created_at": created_at,
            "files": downloadable_files_list
        }

    except sqlite3.Error as e:
        error_logger.error(f"Download Info: SQLite error for state {state_id}, user {user_id}: {e}")
        return None
    except Exception as e:
        error_logger.error(f"Download Info: Unexpected error for state {state_id}, user {user_id}: {e}\n{traceback.format_exc()}")
        return None
    finally:
        if conn:
            conn.close()

def load_state(state_id_to_load: str) -> Optional[MarketIntelligenceState]:
    db_path = get_db_path()
    try:
        conn = sqlite3.connect(db_path)
        cursor_obj = conn.cursor()
        cursor_obj.execute('SELECT state_data FROM states WHERE id = ?', (state_id_to_load,))
        result_data_row = cursor_obj.fetchone()
        conn.close()
        if result_data_row:
            loaded_state = MarketIntelligenceState(**json.loads(result_data_row[0]))
            logger.info(f"State loaded: ID={state_id_to_load}, Domain='{loaded_state.market_domain}' from {db_path}")
            return loaded_state
        else:
            logger.warning(f"State ID '{state_id_to_load}' not found in database at {db_path}.")
            return None
    except Exception as e_load_state:
        error_logger.error(f"Failed to load state {state_id_to_load} from {db_path}: {e_load_state}")
        return None

def save_chat_message(session_id_val: str, message_type_val: str, content_val: str):
    db_path = get_db_path()
    try:
        conn = sqlite3.connect(db_path)
        cursor_obj = conn.cursor()
        cursor_obj.execute('INSERT INTO chat_history (session_id, message_type, content, timestamp) VALUES (?, ?, ?, ?)',
                           (session_id_val, message_type_val, content_val, datetime.now()))
        conn.commit()
        conn.close()
        logger.info(f"Chat message saved: SessionID='{session_id_val}', Type='{message_type_val}' to {db_path}")
    except Exception as e_save_chat:
        error_logger.error(f"Failed to save chat message for SessionID '{session_id_val}' to {db_path}: {e_save_chat}")

def load_chat_history(session_id_val: str) -> List[Dict[str, Any]]:
    db_path = get_db_path()
    try:
        conn = sqlite3.connect(db_path)
        cursor_obj = conn.cursor()
        cursor_obj.execute('SELECT message_type, content FROM chat_history WHERE session_id = ? ORDER BY timestamp ASC', (session_id_val,))
        messages_history_list = [{"type": row[0], "content": row[1]} for row in cursor_obj.fetchall()]
        conn.close()
        logger.info(f"Chat history loaded: SessionID='{session_id_val}', Messages Count={len(messages_history_list)} from {db_path}")
        return messages_history_list
    except Exception as e_load_chat:
        error_logger.error(f"Failed to load chat history for SessionID '{session_id_val}' from {db_path}: {e_load_chat}")
        return []

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def search_with_tavily(search_query: str, user_id: Optional[str] = None) -> List[str]:
    cache_user_prefix = user_id if user_id else "global_fallback"
    normalized_cache_key = f"tavily_search_{cache_user_prefix}_{search_query.lower().replace(' ', '_')}"
    if normalized_cache_key in search_results_cache:
        logger.info(f"Tavily Search: Cache hit for query: '{search_query}', UserID: {user_id or 'N/A (global_fallback used)'}")
        return search_results_cache[normalized_cache_key]
    tavily_api_key_val = get_api_key("TAVILY", user_id=user_id)
    if not tavily_api_key_val:
        error_logger.error(f"Tavily Search: TAVILY_API_KEY not found for UserID: {user_id or 'N/A (global_fallback attempted)'}. Cannot perform search.")
        raise ValueError(f"Tavily API key not available for UserID: {user_id or 'N/A (global_fallback attempted)'}. Search cannot proceed.")
    try:
        logger.info(f"Tavily Search: Performing API search for query: '{search_query}', UserID: {user_id or 'N/A'}.")
        response = requests.post("https://api.tavily.com/search", headers={"Content-Type": "application/json", "Accept": "application/json"},
                                 json={"api_key": tavily_api_key_val, "query": search_query, "search_depth": "advanced", "include_answer": False, "max_results": 7})
        response.raise_for_status()
        response_data = response.json()
        extracted_urls = [r["url"] for r in response_data.get("results", []) if r.get("url")]
        search_results_cache[normalized_cache_key] = extracted_urls
        logger.info(f"Tavily Search: Retrieved {len(extracted_urls)} URLs for query: '{search_query}', UserID: {user_id or 'N/A'}")
        return extracted_urls
    except requests.exceptions.HTTPError as e_tavily_http:
        error_logger.error(f"Tavily Search API HTTP error for query '{search_query}', UserID: {user_id or 'N/A'}: {e_tavily_http}. Response: {e_tavily_http.response.text[:200]}")
        raise
    except requests.exceptions.RequestException as e_tavily_req:
        error_logger.error(f"Tavily Search API request failed (network/other) for query '{search_query}', UserID: {user_id or 'N/A'}: {e_tavily_req}")
        raise
    except Exception as e_tavily_other:
        error_logger.error(f"Unexpected error during Tavily search for query '{search_query}', UserID: {user_id or 'N/A'}: {e_tavily_other}\n{traceback.format_exc()}")
        raise

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def search_with_serpapi(search_query: str, user_id: Optional[str] = None) -> List[str]:
    if SerpApiClient is None:
        logger.warning("SerpAPI search called but library not available. Skipping.")
        return []
    cache_user_prefix = user_id if user_id else "global_fallback"
    normalized_cache_key = f"serpapi_search_{cache_user_prefix}_{search_query.lower().replace(' ', '_')}"
    if normalized_cache_key in search_results_cache:
        logger.info(f"SerpAPI Search: Cache hit for query: '{search_query}', UserID: {user_id or 'N/A (global_fallback used)'}")
        return search_results_cache[normalized_cache_key]
    api_key = get_api_key("SERPAPI", user_id=user_id)
    if not api_key:
        logger.warning(f"SerpAPI Search: SERPAPI_API_KEY not found for UserID: {user_id or 'N/A (global_fallback attempted)'}. Skipping SerpAPI search.")
        return [] # Original behavior was to return empty list, not raise error.

    params = {
        "q": search_query,
        "api_key": api_key,
        "engine": "google", # Or other engines if configurable
        "num": 10 # Number of results
    }

    try:
        logger.info(f"SerpAPI Search: Performing API search for query: '{search_query}', UserID: {user_id or 'N/A'}")
        search = SerpApiClient(params)
        results = search.get_dict()
        extracted_urls = [r["link"] for r in results.get("organic_results", []) if "link" in r]
        search_results_cache[normalized_cache_key] = extracted_urls
        logger.info(f"SerpAPI Search: Retrieved {len(extracted_urls)} URLs for query: '{search_query}', UserID: {user_id or 'N/A'}")
        return extracted_urls
    except requests.exceptions.RequestException as e_serpapi_req:
        error_logger.error(f"SerpAPI Search API request failed for query '{search_query}', UserID: {user_id or 'N/A'}: {e_serpapi_req}")
        return []
    except Exception as e_serpapi_other:
        error_logger.error(f"Unexpected error during SerpAPI search for query '{search_query}', UserID: {user_id or 'N/A'}: {e_serpapi_other}\n{traceback.format_exc()}")
        return []

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def fetch_from_newsapi_direct(query: str, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
    if NewsApiClient is None:
        logger.warning("NewsAPI direct search called but NewsApiClient library not available. Skipping.")
        return []
    cache_user_prefix = user_id if user_id else "global_fallback"
    cache_key = f"newsapi_direct_search_{cache_user_prefix}_{query.lower().replace(' ', '_')}"
    if cache_key in search_results_cache:
        logger.info(f"NewsAPI Direct: Cache hit for query: '{query}', UserID: {user_id or 'N/A (global_fallback used)'}")
        return search_results_cache[cache_key]
    api_key = get_api_key("NEWS_API", user_id=user_id)
    if not api_key:
        logger.warning(f"NewsAPI Direct: NEWSAPI_API_KEY not found for UserID: {user_id or 'N/A (global_fallback attempted)'}. Skipping NewsAPI direct search.")
        return []
    transformed_articles = []
    try:
        logger.info(f"NewsAPI Direct: Performing API search for query: '{query}', UserID: {user_id or 'N/A'}")
        newsapi = NewsApiClient(api_key=api_key)
        all_articles_raw = newsapi.get_everything(q=query, language='en', sort_by='relevancy', page_size=10)
        for article in all_articles_raw.get('articles', []):
            transformed_articles.append({"source": "NewsAPI - " + article.get('source', {}).get('name', 'Unknown Source'),
                                         "title": article.get('title'), "summary": article.get('description'),
                                         "full_content": article.get('content', article.get('description')),
                                         "url": article.get('url'), "publishedAt": article.get('publishedAt')})
        search_results_cache[cache_key] = transformed_articles
        logger.info(f"NewsAPI Direct: Retrieved {len(transformed_articles)} articles for query: '{query}', UserID: {user_id or 'N/A'}")
        return transformed_articles
    except Exception as e_newsapi:
        error_logger.error(f"NewsAPI Direct search failed for query '{query}', UserID: {user_id or 'N/A'}: {e_newsapi}\n{traceback.format_exc()}")
        return []

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def fetch_from_mediastack_direct(query: str, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
    cache_user_prefix = user_id if user_id else "global_fallback"
    cache_key = f"mediastack_direct_search_{cache_user_prefix}_{query.lower().replace(' ', '_')}"
    if cache_key in search_results_cache:
        logger.info(f"MediaStack Direct: Cache hit for query: '{query}', UserID: {user_id or 'N/A (global_fallback used)'}")
        return search_results_cache[cache_key]
    api_key = get_api_key("MEDIASTACK", user_id=user_id)
    if not api_key:
        logger.warning(f"MediaStack Direct: MEDIASTACK_API_KEY not found for UserID: {user_id or 'N/A (global_fallback attempted)'}. Skipping MediaStack direct search.")
        return []
    endpoint = "http://api.mediastack.com/v1/news"
    params = {'access_key': api_key, 'keywords': query, 'limit': 10, 'languages': 'en'}
    transformed_articles = []
    try:
        logger.info(f"MediaStack Direct: Performing API search for query: '{query}', UserID: {user_id or 'N/A'}")
        response = requests.get(endpoint, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        for article in data.get('data', []):
            transformed_articles.append({"source": "MediaStack - " + str(article.get('source', 'Unknown Source')),
                                         "title": article.get('title'), "summary": article.get('description'),
                                         "full_content": article.get('description'), "url": article.get('url'),
                                         "publishedAt": article.get('published_at')})
        search_results_cache[cache_key] = transformed_articles
        logger.info(f"MediaStack Direct: Retrieved {len(transformed_articles)} articles for query: '{query}', UserID: {user_id or 'N/A'}")
        return transformed_articles
    except requests.exceptions.HTTPError as e_http:
        error_logger.error(f"MediaStack Direct API HTTP error for query '{query}', UserID: {user_id or 'N/A'}: {e_http}. Response: {e_http.response.text[:200]}")
        return []
    except requests.exceptions.RequestException as e_mediastack_req:
        error_logger.error(f"MediaStack Direct API request failed for query '{query}', UserID: {user_id or 'N/A'}: {e_mediastack_req}")
        return []
    except Exception as e_mediastack_other:
        error_logger.error(f"Unexpected error during MediaStack Direct search for query '{query}', UserID: {user_id or 'N/A'}: {e_mediastack_other}\n{traceback.format_exc()}")
        return []

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def fetch_financial_data_fmp(query: str, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
    if fmpsdk is None:
        logger.warning("FMP SDK called but library not available.")
        return []
    cache_user_prefix = user_id if user_id else "global_fallback"
    cache_key = f"fmp_data_{cache_user_prefix}_{query.lower().replace(' ', '_')}"
    if cache_key in search_results_cache:
        logger.info(f"FMP: Cache hit for query: '{query}', UserID: {user_id or 'N/A'}")
        return search_results_cache[cache_key]

    api_key = get_api_key("FINANCIAL_MODELING_PREP")
    if not api_key:
        logger.warning(f"FMP_API_KEY not found for UserID: {user_id or 'N/A'}. Skipping FMP data fetch.")
        return []
    potential_symbols = re.findall(r'\b([A-Z]{1,5})\b', query)
    symbol_to_use = potential_symbols[0] if potential_symbols else None
    if not symbol_to_use:
        logger.warning(f"No valid stock symbol found in query '{query}'. Skipping FMP data fetch.")
        return []
    fetched_fmp_data = []
    try:
        logger.info(f"FMP: Fetching data for symbol '{symbol_to_use}'")
        profile = fmpsdk.company_profile(apikey=api_key, symbol=symbol_to_use)
        quote = fmpsdk.quote(apikey=api_key, symbol=symbol_to_use)
        if profile:
            fetched_fmp_data.append({"source": "FinancialModelingPrep", "type": "company_profile", "symbol": symbol_to_use, "data": profile[0] if isinstance(profile, list) else profile})
        if quote:
            fetched_fmp_data.append({"source": "FinancialModelingPrep", "type": "stock_quote", "symbol": symbol_to_use, "data": quote[0] if isinstance(quote, list) else quote})
        logger.info(f"FMP: Fetched {len(fetched_fmp_data)} data points for symbol {symbol_to_use}.")
        search_results_cache[cache_key] = fetched_fmp_data
        return fetched_fmp_data
    except Exception as e:
        error_logger.error(f"FMP data fetching failed for symbol {symbol_to_use}: {e}\n{traceback.format_exc()}")
        return []

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=8))
def fetch_financial_data_alphavantage(query: str, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
    if TimeSeries is None or FundamentalData is None:
        logger.warning("Alpha Vantage library called but not fully available.")
        return []
    cache_user_prefix = user_id if user_id else "global_fallback"
    cache_key = f"alphavantage_data_{cache_user_prefix}_{query.lower().replace(' ', '_')}"
    if cache_key in search_results_cache:
        logger.info(f"AlphaVantage: Cache hit for query: '{query}', UserID: {user_id or 'N/A'}")
        return search_results_cache[cache_key]

    api_key = get_api_key("ALPHA_VANTAGE")
    if not api_key:
        logger.warning(f"ALPHA_VANTAGE_API_KEY not found for UserID: {user_id or 'N/A'}. Skipping Alpha Vantage data fetch.")
        return []
    potential_symbols = re.findall(r'\b([A-Z]{1,5})\b', query)
    symbol_to_use = potential_symbols[0] if potential_symbols else None
    if not symbol_to_use:
        logger.warning(f"No valid stock symbol found in query '{query}'. Skipping Alpha Vantage data fetch.")
        return []
    fetched_av_data = []
    try:
        if TimeSeries:
            ts = TimeSeries(key=api_key, output_format='json')
            data_ts, meta_data_ts = ts.get_daily(symbol=symbol_to_use, outputsize='compact')
            if data_ts:
                latest_date = sorted(data_ts.keys(), reverse=True)[0]
                latest_data_point = data_ts[latest_date]
                fetched_av_data.append({"source": "AlphaVantage", "type": "daily_time_series_latest", "symbol": symbol_to_use, "data": {"date": latest_date, **latest_data_point}})
        if FundamentalData:
            fd = FundamentalData(key=api_key, output_format='json')
            try:
                data_overview, _ = fd.get_company_overview(symbol=symbol_to_use)
                if data_overview:
                    fetched_av_data.append({"source": "AlphaVantage", "type": "company_overview", "symbol": symbol_to_use, "data": data_overview})
            except Exception as e_overview:
                logger.warning(f"AlphaVantage: Could not fetch company overview for {symbol_to_use}: {e_overview}")
        logger.info(f"AlphaVantage: Fetched {len(fetched_av_data)} data points for symbol {symbol_to_use}.")
        search_results_cache[cache_key] = fetched_av_data
        return fetched_av_data
    except Exception as e:
        error_logger.error(f"AlphaVantage data fetching failed for symbol {symbol_to_use}: {e}\n{traceback.format_exc()}")
        return []

async def fetch_url_content(url_to_fetch: str) -> Dict[str, Any]:
    try:
        logger.info(f"WebBaseLoader: Loading content from URL: {url_to_fetch}")
        loader = WebBaseLoader([url_to_fetch])
        loaded_docs = await loader.aload()
        doc_object = loaded_docs[0] if loaded_docs else None
        if doc_object:
            raw_page_content = doc_object.page_content
            cleaned_page_content = re.sub(r'\n\s*\n', '\n\n', raw_page_content).strip()
            summary_text = cleaned_page_content[:1000]
            document_title = doc_object.metadata.get("title", "") or os.path.basename(url_to_fetch)
            if not document_title:
                document_title = "Untitled Document"
            logger.info(f"WebBaseLoader: Loaded from {url_to_fetch}. Title: '{document_title}'. Summary (first 50): '{summary_text[:50]}...' ")
            return {"source": url_to_fetch, "title": document_title, "summary": summary_text, "full_content": cleaned_page_content, "url": url_to_fetch}
        else:
            logger.warning(f"WebBaseLoader: No document object returned from {url_to_fetch}")
            return {"source": url_to_fetch, "title": "Content Not Loaded", "summary": "", "full_content": "", "url": url_to_fetch}
    except Exception as e_fetch_url:
        error_logger.error(f"WebBaseLoader: Failed to load content from URL '{url_to_fetch}': {e_fetch_url}\n{traceback.format_exc()}")
        return {"source": url_to_fetch, "title": f"Failed to Load: {os.path.basename(url_to_fetch)}", "summary": str(e_fetch_url), "full_content": "", "url": url_to_fetch}

def get_agent_base_reports_dir():
    agent_script_dir = os.path.dirname(os.path.abspath(__file__))
    base_reports_dir = os.path.join(agent_script_dir, "reports1")
    if os.environ.get("VERCEL_ENV"):
        base_reports_dir = os.path.join("/tmp", "reports1")
        logger.info(f"Vercel environment detected. Using /tmp/reports1 for reports base.")
    else:
        logger.info(f"Local environment. Using {base_reports_dir} for reports base.")
    os.makedirs(base_reports_dir, exist_ok=True)
    return base_reports_dir

async def market_data_collector(current_state: MarketIntelligenceState, uploaded_file_paths: Optional[List[str]] = None) -> Dict[str, Any]:
    logger.info(f"Market Data Collector: Domain='{current_state.market_domain}', Query='{current_state.query or 'N/A'}'")
    ts_string = datetime.now().strftime("%Y%m%d_%H%M%S")
    query_prefix = re.sub(r'[^a-zA-Z0-9_-]', '_', (current_state.query or "general").lower().replace(' ', '_')[:20])
    base_reports_path = get_agent_base_reports_dir()
    run_report_dir = os.path.join(base_reports_path, f"{query_prefix}_{ts_string}")
    try:
        os.makedirs(run_report_dir, exist_ok=True)
        current_state.report_dir = run_report_dir
        logger.info(f"Market Data Collector: Report directory set to: {run_report_dir}")
    except Exception as e_mkdir_report:
        error_logger.critical(f"CRITICAL: Failed to create report directory '{run_report_dir}': {e_mkdir_report}")
        raise IOError(f"Cannot create report directory '{run_report_dir}': {e_mkdir_report}")

    all_fetched_data = [] # Initialize here

    # Process uploaded document IDs first
    if current_state.uploaded_document_ids and current_state.user_id:
        logger.info(f"Market Data Collector: Processing {len(current_state.uploaded_document_ids)} uploaded document IDs.")
        supabase = get_supabase_client() # Get Supabase client instance
        if supabase:
            for doc_id in current_state.uploaded_document_ids:
                try:
                    # Fetch only necessary fields, ensure uploader_id matches for security if needed,
                    # but agent_logic typically runs with service_role_key so it has broad access.
                    # Assuming user_id in current_state is the owner.
                    doc_response = supabase.table("documents").select("original_filename, text, file_extension").eq("id", doc_id).eq("uploader_id", current_state.user_id).maybe_single().execute()

                    if doc_response.data and doc_response.data.get("text"):
                        doc_content = doc_response.data["text"]
                        doc_title = doc_response.data.get("original_filename", f"Document {doc_id}")
                        # Create a SyncedArticle-like structure or directly a dict for all_fetched_data
                        all_fetched_data.append({
                            "source": f"document_id://{doc_id}", # Indicate source is an uploaded document
                            "title": doc_title,
                            "summary": doc_content[:500] + "..." if doc_content and len(doc_content) > 500 else doc_content, # Basic summary
                            "full_content": doc_content,
                            "url": f"app://document/{doc_id}" # Placeholder URL scheme
                        })
                        logger.info(f"Market Data Collector: Added content from uploaded document ID {doc_id}.")
                    else:
                        logger.warning(f"Market Data Collector: Document ID {doc_id} not found or text is empty for user {current_state.user_id}.")
                except Exception as e_fetch_doc:
                    error_logger.error(f"Market Data Collector: Failed to fetch/process uploaded document ID {doc_id}: {e_fetch_doc}")
        else:
            logger.warning("Market Data Collector: Supabase client not available. Skipping processing of uploaded_document_ids.")

    current_state.uploaded_files_content = [data for data in all_fetched_data if data["source"].startswith("document_id://")]


    json_file_path = os.path.join(run_report_dir, f"{current_state.market_domain.lower().replace(' ', '_')}_data_sources.json")
    csv_file_path = os.path.join(run_report_dir, f"{current_state.market_domain.lower().replace(' ', '_')}_data_sources.csv")
    news_search_query = f"{current_state.query} {current_state.market_domain} news trends developments emerging technologies"
    competitor_search_query = f"{current_state.query} {current_state.market_domain} competitor landscape key players market share"
    news_urls_list = []
    competitor_urls_list = []
    serpapi_news_urls_list = []
    serpapi_competitor_urls_list = []

    try:
        logger.info("Attempting Tavily search for news URLs...")
        news_urls_list = search_with_tavily(news_search_query)
        logger.info(f"Tavily news search returned {len(news_urls_list)} URLs.")
    except Exception as e_tavily_news:
        error_logger.error(f"Tavily news search failed: {e_tavily_news}")
    try:
        logger.info("Attempting Tavily search for competitor URLs...")
        competitor_urls_list = search_with_tavily(competitor_search_query)
        logger.info(f"Tavily competitor search returned {len(competitor_urls_list)} URLs.")
    except Exception as e_tavily_comp:
        error_logger.error(f"Tavily competitor search failed: {e_tavily_comp}")
    if SerpApiClient is not None:
        try:
            logger.info("Attempting SerpAPI search for news URLs...")
            serpapi_news_urls_list = search_with_serpapi(news_search_query)
            logger.info(f"SerpAPI news search returned {len(serpapi_news_urls_list)} URLs.")
        except Exception as e_serp_news:
            error_logger.error(f"SerpAPI news search failed: {e_serp_news}")
        try:
            logger.info("Attempting SerpAPI search for competitor URLs...")
            serpapi_competitor_urls_list = search_with_serpapi(competitor_search_query)
            logger.info(f"SerpAPI competitor search returned {len(serpapi_competitor_urls_list)} URLs.")
        except Exception as e_serp_comp:
            error_logger.error(f"SerpAPI competitor search failed: {e_serp_comp}")
    else:
        logger.info("SerpAPI library not available, skipping SerpAPI searches.")
    combined_unique_urls = list(set(news_urls_list + competitor_urls_list + serpapi_news_urls_list + serpapi_competitor_urls_list))
    logger.info(f"Market Data Collector: Total unique URLs to process: {len(combined_unique_urls)}")
    all_fetched_data = []
    current_query_or_domain = current_state.query if current_state.query else current_state.market_domain
    if NewsApiClient is not None:
        try:
            logger.info(f"Fetching from NewsAPI for query: '{current_query_or_domain}'")
            newsapi_articles = fetch_from_newsapi_direct(current_query_or_domain)
            all_fetched_data.extend(newsapi_articles)
            logger.info(f"Retrieved {len(newsapi_articles)} articles from NewsAPI.")
        except Exception as e_newsapi:
            error_logger.error(f"Failed to fetch from NewsAPI: {e_newsapi}")
    try:
        logger.info(f"Fetching from MediaStack for query: '{current_query_or_domain}'")
        mediastack_articles = fetch_from_mediastack_direct(current_query_or_domain)
        all_fetched_data.extend(mediastack_articles)
        logger.info(f"Retrieved {len(mediastack_articles)} articles from MediaStack.")
    except Exception as e_mstack:
        error_logger.error(f"Failed to fetch from MediaStack: {e_mstack}")
    current_state.financial_data = []
    if fmpsdk is not None:
        try:
            logger.info(f"Fetching financial data from FMP for query: '{current_query_or_domain}'")
            fmp_fin_data = fetch_financial_data_fmp(current_query_or_domain)
            current_state.financial_data.extend(fmp_fin_data)
            logger.info(f"Retrieved {len(fmp_fin_data)} data items from FMP.")
        except Exception as e_fmp_fin:
            error_logger.error(f"Failed to fetch financial data from FMP: {e_fmp_fin}")
    if TimeSeries is not None and FundamentalData is not None:
        try:
            logger.info(f"Fetching financial data from Alpha Vantage for query: '{current_query_or_domain}'")
            av_fin_data = fetch_financial_data_alphavantage(current_query_or_domain)
            current_state.financial_data.extend(av_fin_data)
            logger.info(f"Retrieved {len(av_fin_data)} data items from Alpha Vantage.")
        except Exception as e_av_fin:
            error_logger.error(f"Failed to fetch financial data from Alpha Vantage: {e_av_fin}")
    logger.info(f"Total financial data items collected: {len(current_state.financial_data)}")
    for idx, loop_url in enumerate(combined_unique_urls):
        if any(article.get("url") == loop_url for article in all_fetched_data):
            logger.info(f"Market Data Collector: Skipping URL {loop_url} as it was fetched directly.")
            continue
        logger.info(f"Market Data Collector: Processing URL {idx + 1}/{len(combined_unique_urls)}: {loop_url}")
        content_data = await fetch_url_content(loop_url)
        all_fetched_data.append(content_data)

    current_state.raw_news_data = all_fetched_data
    current_state.competitor_data = all_fetched_data

    try:
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(all_fetched_data, f, indent=4)
        logger.info(f"Market Data Collector: Data saved to JSON: {json_file_path}")
        current_state.download_files["raw_data_json"] = json_file_path
    except Exception as e_json:
        error_logger.error(f"Failed to save JSON '{json_file_path}': {e_json}")
    try:
        with open(csv_file_path, "w", newline="", encoding="utf-8") as f:
            field_names_csv = ["title", "summary", "url", "source", "full_content"]
            writer_csv = csv.DictWriter(f, fieldnames=field_names_csv, extrasaction="ignore")
            writer_csv.writeheader()
            writer_csv.writerows(all_fetched_data)
        logger.info(f"Market Data Collector: Data saved to CSV: {csv_file_path}")
        current_state.download_files["raw_data_csv"] = csv_file_path
    except Exception as e_csv:
        error_logger.error(f"Failed to save CSV '{csv_file_path}': {e_csv}")
    save_state(current_state)
    logger.info("Market Data Collector: Node completed.")
    return current_state.model_dump()

def llm_json_parser_robust(llm_output_str: str, default_return_val: Any = None) -> Any:
    logger.debug(f"LLM JSON Parser: Attempting to parse: {llm_output_str[:200]}...")
    try:
        cleaned_llm_output = re.sub(r"\`\`\`json\s*([\s\S]*?)\s*\`\`\`", r"\1", llm_output_str.strip(), flags=re.IGNORECASE)
        start_brace = cleaned_llm_output.find('{')
        start_bracket = cleaned_llm_output.find('[')
        if start_brace == -1 and start_bracket == -1:
            logger.warning(f"LLM JSON Parser: No JSON object/array start found. Output: {cleaned_llm_output[:200]}")
            return default_return_val if default_return_val is not None else []
        json_start_char = '{' if (start_brace != -1 and (start_bracket == -1 or start_brace < start_bracket)) else '['
        json_start_index = start_brace if json_start_char == '{' else start_bracket
        open_count = 0
        json_end_index = -1
        for i in range(json_start_index, len(cleaned_llm_output)):
            if cleaned_llm_output[i] == json_start_char:
                open_count += 1
            elif (json_start_char == '{' and cleaned_llm_output[i] == '}') or \
                 (json_start_char == '[' and cleaned_llm_output[i] == ']'):
                open_count -= 1
            if open_count == 0:
                json_end_index = i
                break
        if json_end_index == -1:
            logger.warning(f"LLM JSON Parser: Could not find matching end for '{json_start_char}'. Output: {cleaned_llm_output[:200]}")
            return default_return_val if default_return_val is not None else []
        json_str_to_parse = cleaned_llm_output[json_start_index:json_end_index + 1]
        parsed_json = json.loads(json_str_to_parse)
        logger.debug("LLM JSON Parser: Successfully parsed JSON.")
        return parsed_json
    except json.JSONDecodeError as e_json_decode:
        error_logger.warning(f"LLM JSON Parser: Parsing failed: {e_json_decode}. String attempted: '{json_str_to_parse[:500] if 'json_str_to_parse' in locals() else cleaned_llm_output[:500]}'")
        return default_return_val if default_return_val is not None else []

async def trend_analyzer(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Trend Analyzer: Domain='{current_state.market_domain}', UserID='{current_state.user_id or 'N/A'}'")
    default_trends_list = [{"trend_name": "Default Trend", "description": "No specific trends identified.", "supporting_evidence": "N/A", "estimated_impact": "Unknown", "timeframe": "Unknown"}]
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.2
        if user_google_api_key:
            logger.info(f"Trend Analyzer: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Trend Analyzer: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Trend Analyzer: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        prompt = ChatPromptTemplate.from_messages([("system", "You are an expert market analyst for {market_domain}. Identify key trends from the provided data. Return a JSON array of objects, each with 'trend_name' (string), 'description' (string), 'supporting_evidence' (string, cite sources if possible), 'estimated_impact' ('High'/'Medium'/'Low'), 'timeframe' ('Short-term'/'Medium-term'/'Long-term'). Aim for 3-5 trends."),
                                                 ("human", "Data for {market_domain} (Query: {query}):\n\nNews/Competitor Info (sample):\n{input_json_data}")])
        chain = prompt | llm | StrOutputParser()
        limited_news_data = current_state.raw_news_data[:5] if current_state.raw_news_data else []
        limited_competitor_data = current_state.competitor_data[:5] if current_state.competitor_data else []
        input_data_for_llm = {"news_sample": limited_news_data, "competitors_sample": limited_competitor_data}
        logger.info(f"Trend Analyzer: Invoking LLM for state {current_state.state_id}. News items: {len(limited_news_data)}, Competitor items: {len(limited_competitor_data)}")
        llm_output_string = await chain.ainvoke({"market_domain": current_state.market_domain, "query": current_state.query or "general", "input_json_data": json.dumps(input_data_for_llm)})
        parsed_trends = llm_json_parser_robust(llm_output_string, default_return_val=default_trends_list)

        # Validate structure: list of dicts
        if isinstance(parsed_trends, list) and all(isinstance(t, dict) for t in parsed_trends):
            current_state.market_trends = parsed_trends
            current_state.num_trends = len(parsed_trends)
            logger.info(f"Trend Analyzer: Successfully parsed and stored {current_state.num_trends} trends for state {current_state.state_id}.")
        else:
            logger.warning(f"Trend Analyzer: Parsed trends not a list of dicts for state {current_state.state_id}. Using default. LLM Output: {llm_output_string[:200]}")
            current_state.market_trends = default_trends_list
            current_state.num_trends = len(default_trends_list)
            current_state.node_errors.append({
                "node": "trend_analyzer",
                "error": "Parsed trends not a list of dicts or LLM output parsing failed.",
                "details": llm_output_string[:200] # Log snippet of problematic output
            })

        if current_state.report_dir:
            trends_json_path = os.path.join(current_state.report_dir, "market_trends.json")
            try:
                with open(trends_json_path, "w", encoding="utf-8") as f:
                    json.dump(parsed_trends, f, indent=4)
                current_state.download_files["trends_json"] = trends_json_path
                logger.info(f"Trend Analyzer: Saved trends to {trends_json_path} for state {current_state.state_id}")
            except Exception as e_json:
                error_logger.error(f"Trend Analyzer: Failed to save trends JSON '{trends_json_path}' for state {current_state.state_id}: {e_json}")
        else:
            logger.warning(f"Trend Analyzer: current_state.report_dir not set for state {current_state.state_id}. Cannot save trends JSON.")
    except ValueError as ve:
        error_logger.error(f"Trend Analyzer: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.market_trends = default_trends_list
        current_state.num_trends = len(default_trends_list)
        current_state.node_errors.append({"node": "trend_analyzer", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e_trend:
        error_logger.error(f"Trend Analyzer: Failed for state {current_state.state_id} ('{current_state.market_domain}'): {e_trend}\n{traceback.format_exc()}")
        current_state.market_trends = default_trends_list
        current_state.num_trends = len(default_trends_list)
        current_state.node_errors.append({"node": "trend_analyzer", "error": f"Exception: {str(e_trend)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"Trend Analyzer: Node completed for state {current_state.state_id}.")
    return current_state.model_dump()

async def opportunity_identifier(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Opportunity Identifier: Domain='{current_state.market_domain}', UserID='{current_state.user_id or 'N/A'}'")
    default_ops = [{"opportunity_name": "Default Opportunity", "description": "N/A", "target_segment": "N/A", "competitive_advantage": "N/A", "estimated_potential": "Unknown", "timeframe_to_capture": "Unknown"}]
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.3
        if user_google_api_key:
            logger.info(f"Opportunity Identifier: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Opportunity Identifier: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Opportunity Identifier: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        prompt = ChatPromptTemplate.from_messages([("system", "Identify market opportunities for {market_domain} based on trends, news, and competitor data. Return JSON array: 'opportunity_name', 'description', 'target_segment', 'competitive_advantage', 'estimated_potential' (High/Medium/Low), 'timeframe_to_capture'. Min 2-3."),
                                                 ("human", "Context for {market_domain}:\nTrends: {trends_json}\nNews/Competitors (sample): {data_json}")])
        chain = prompt | llm | StrOutputParser()
        limited_news = current_state.raw_news_data[:5] if current_state.raw_news_data else []
        # Make sure default_ops has all keys the prompt expects if parsed_ops is not a list of dicts later
        llm_output = await chain.ainvoke({
            "market_domain": current_state.market_domain,
            "trends_json": json.dumps(current_state.market_trends[:5] if current_state.market_trends else []), # Ensure market_trends exists
            "data_json": json.dumps({"news_sample": limited_news})
        })
        
        parsed_ops = llm_json_parser_robust(llm_output, default_return_val=default_ops)

        # Validate structure: list of dicts
        if isinstance(parsed_ops, list) and all(isinstance(op, dict) for op in parsed_ops):
            current_state.opportunities = parsed_ops
            current_state.num_opportunities = len(parsed_ops)
            logger.info(f"Opportunity Identifier: Successfully parsed and stored {current_state.num_opportunities} opportunities for state {current_state.state_id}.")
        else:
            logger.warning(f"Opportunity Identifier: Parsed opportunities not a list of dicts for state {current_state.state_id}. Using default. LLM Output: {llm_output[:200]}")
            current_state.opportunities = default_ops
            current_state.num_opportunities = len(default_ops)
            current_state.node_errors.append({
                "node": "opportunity_identifier",
                "error": "Parsed opportunities not a list of dicts or LLM output parsing failed.",
                "details": llm_output[:200]
            })

        if current_state.report_dir:
            opportunities_json_path = os.path.join(current_state.report_dir, "opportunities.json")
            try:
                with open(opportunities_json_path, "w", encoding="utf-8") as f:
                    json.dump(parsed_ops, f, indent=4)
                current_state.download_files["opportunities_json"] = opportunities_json_path
                logger.info(f"Opportunity Identifier: Saved opportunities to {opportunities_json_path} for state {current_state.state_id}")
            except Exception as e_json:
                error_logger.error(f"Opportunity Identifier: Failed to save opportunities JSON '{opportunities_json_path}' for state {current_state.state_id}: {e_json}")
        else:
            logger.warning(f"Opportunity Identifier: current_state.report_dir not set for state {current_state.state_id}. Cannot save opportunities JSON.")
    except ValueError as ve:
        error_logger.error(f"Opportunity Identifier: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.opportunities = default_ops
        current_state.num_opportunities = len(default_ops)
        current_state.node_errors.append({"node": "opportunity_identifier", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e:
        error_logger.error(f"Opportunity Identifier: Failed for state {current_state.state_id}: {e}\n{traceback.format_exc()}")
        current_state.opportunities = default_ops
        current_state.num_opportunities = len(default_ops)
        current_state.node_errors.append({"node": "opportunity_identifier", "error": f"Exception: {str(e)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"Opportunity Identifier: Found {len(current_state.opportunities)} opportunities for state {current_state.state_id}.")
    return current_state.model_dump()

async def strategy_recommender(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Strategy Recommender: Domain='{current_state.market_domain}', UserID='{current_state.user_id or 'N/A'}'")
    default_strats = [{"strategy_title": "Default Strategy", "description": "N/A", "implementation_steps": [], "expected_outcome": "N/A", "resource_requirements": "N/A", "priority_level": "Unknown", "success_metrics": "N/A"}]
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.3
        if user_google_api_key:
            logger.info(f"Strategy Recommender: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Strategy Recommender: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Strategy Recommender: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        prompt = ChatPromptTemplate.from_messages([("system", "Recommend strategies for {market_domain} based on opportunities, trends, and competitor data. Return JSON array: 'strategy_title', 'description', 'implementation_steps' (list), 'expected_outcome', 'resource_requirements', 'priority_level', 'success_metrics'. Min 2-3."),
                                                 ("human", "Context for {market_domain}:\nOpportunities: {ops_json}\nTrends: {trends_json}\nCompetitors (sample): {comp_json}")])
        chain = prompt | llm | StrOutputParser()
        limited_comp = current_state.competitor_data[:5] if current_state.competitor_data else []
        llm_output = await chain.ainvoke({
            "market_domain": current_state.market_domain,
            "ops_json": json.dumps(current_state.opportunities[:5] if current_state.opportunities else []), # Ensure opportunities exist
            "trends_json": json.dumps(current_state.market_trends[:5] if current_state.market_trends else []), # Ensure market_trends exist
            "comp_json": json.dumps({"competitors_sample": limited_comp})
        })

        parsed_strats = llm_json_parser_robust(llm_output, default_return_val=default_strats)
        if not isinstance(parsed_strats, list) or not all(isinstance(s, dict) for s in parsed_strats):
            logger.warning(f"Strategy Recommender: Parsed strategies not a list of dicts for state {current_state.state_id}. Using default. Output: {llm_output[:200]}")
            current_state.strategic_recommendations = default_strats
            current_state.node_errors.append({
                "node": "strategy_recommender",
                "error": "Parsed strategies not a list of dicts or LLM output parsing failed.",
                "details": llm_output[:200]
            })
        else:
            current_state.strategic_recommendations = parsed_strats
            logger.info(f"Strategy Recommender: Successfully parsed {len(parsed_strats)} strategies for state {current_state.state_id}.")

        if current_state.report_dir:
            strategies_json_path = os.path.join(current_state.report_dir, "strategies.json")
            try:
                with open(strategies_json_path, "w", encoding="utf-8") as f:
                    json.dump(parsed_strats, f, indent=4)
                current_state.download_files["strategies_json"] = strategies_json_path
                logger.info(f"Strategy Recommender: Saved strategies to {strategies_json_path} for state {current_state.state_id}")
            except Exception as e_json:
                error_logger.error(f"Strategy Recommender: Failed to save strategies JSON '{strategies_json_path}' for state {current_state.state_id}: {e_json}")
        else:
            logger.warning(f"Strategy Recommender: current_state.report_dir not set for state {current_state.state_id}. Cannot save strategies JSON.")
    except ValueError as ve:
        error_logger.error(f"Strategy Recommender: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.strategic_recommendations = default_strats
        current_state.node_errors.append({"node": "strategy_recommender", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e:
        error_logger.error(f"Strategy Recommender: Failed for state {current_state.state_id}: {e}\n{traceback.format_exc()}")
        current_state.strategic_recommendations = default_strats
        current_state.node_errors.append({"node": "strategy_recommender", "error": f"Exception: {str(e)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"Strategy Recommender: Generated {len(current_state.strategic_recommendations)} strategies for state {current_state.state_id}.")
    return current_state.model_dump()

async def customer_insights_generator(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Customer Insights Generator: Domain='{current_state.market_domain}', UserID='{current_state.user_id or 'N/A'}'")
    default_insights = [{"segment_name": "Enterprise", "description": "Large organizations with complex needs", "percentage": 35, "key_characteristics": ["High budget", "Long sales cycle", "Multiple stakeholders"], "pain_points": ["Integration complexity", "Security concerns", "Compliance requirements"], "growth_potential": "Medium", "satisfaction_score": 7.8, "retention_rate": 85, "acquisition_cost": "High", "lifetime_value": "Very High"},
                        {"segment_name": "SMB", "description": "Small and medium businesses", "percentage": 45, "key_characteristics": ["Price sensitive", "Quick decision making", "Limited resources"], "pain_points": ["Cost concerns", "Ease of implementation", "Limited technical expertise"], "growth_potential": "High", "satisfaction_score": 8.2, "retention_rate": 75, "acquisition_cost": "Medium", "lifetime_value": "Medium"},
                        {"segment_name": "Startups", "description": "Early stage companies with rapid growth", "percentage": 20, "key_characteristics": ["Innovation focused", "Limited budget", "Agile processes"], "pain_points": ["Scalability", "Quick time-to-value", "Flexible pricing models"], "growth_potential": "Very High", "satisfaction_score": 8.5, "retention_rate": 65, "acquisition_cost": "Low", "lifetime_value": "Variable"}]
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.3
        if user_google_api_key:
            logger.info(f"Customer Insights Generator: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Customer Insights Generator: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Customer Insights Generator: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        prompt = ChatPromptTemplate.from_messages([("system", "You are a customer insights expert for {market_domain}. Based on the provided market data, identify key customer segments and their characteristics. Return a JSON array of objects, each with 'segment_name', 'description', 'percentage' (numeric), 'key_characteristics' (array), 'pain_points' (array), 'growth_potential' (string), 'satisfaction_score' (numeric 1-10), 'retention_rate' (numeric percentage), 'acquisition_cost' (string), 'lifetime_value' (string). Aim for 3-5 segments."),
                                                 ("human", "Market data for {market_domain}:\nOpportunities: {ops_json}\nTrends: {trends_json}\nCompetitors: {comp_json}")])
        chain = prompt | llm | StrOutputParser()
        limited_ops = current_state.opportunities[:5] if current_state.opportunities else []
        limited_trends = current_state.market_trends[:5] if current_state.market_trends else []
        limited_comp = current_state.competitor_data[:5] if current_state.competitor_data else []
        llm_output = await chain.ainvoke({"market_domain": current_state.market_domain, "ops_json": json.dumps(limited_ops), "trends_json": json.dumps(limited_trends), "comp_json": json.dumps(limited_comp)})
        parsed_insights = llm_json_parser_robust(llm_output, default_return_val=default_insights)
        if not isinstance(parsed_insights, list) or not all(isinstance(i, dict) for i in parsed_insights):
            logger.warning(f"Customer Insights Generator: Parsed insights not a list of dicts for state {current_state.state_id}. Using default. Output: {llm_output[:200]}")
            current_state.customer_insights = default_insights
            current_state.node_errors.append({
                "node": "customer_insights_generator",
                "error": "Parsed insights not a list of dicts or LLM output parsing failed.",
                "details": llm_output[:200]
            })
        else:
            current_state.customer_insights = parsed_insights
            logger.info(f"Customer Insights Generator: Successfully parsed {len(parsed_insights)} customer segments for state {current_state.state_id}.")

        if current_state.report_dir:
            insights_json_path = os.path.join(current_state.report_dir, "customer_insights.json")
            try:
                with open(insights_json_path, "w", encoding="utf-8") as f:
                    json.dump(parsed_insights, f, indent=4)
                current_state.download_files["customer_insights_json"] = insights_json_path
                logger.info(f"Customer Insights Generator: Saved insights to {insights_json_path} for state {current_state.state_id}")
            except Exception as e_json:
                error_logger.error(f"Customer Insights Generator: Failed to save insights JSON '{insights_json_path}' for state {current_state.state_id}: {e_json}")
        else:
            logger.warning(f"Customer Insights Generator: current_state.report_dir not set for state {current_state.state_id}. Cannot save insights JSON.")
        db_path = get_db_path()
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            cursor_obj = conn.cursor()
            for segment in parsed_insights:
                if isinstance(segment, dict):
                    segment_id = str(uuid4())
                    cursor_obj.execute('INSERT OR REPLACE INTO customer_insights (id, state_id, segment_name, segment_data) VALUES (?, ?, ?, ?)',
                                       (segment_id, current_state.state_id, segment.get('segment_name', 'Unknown'), json.dumps(segment)))
                else:
                    logger.warning(f"Customer Insights Generator: Skipping invalid segment data during DB save for state {current_state.state_id}: {segment}")
            conn.commit()
        except sqlite3.Error as e_db_sqlite:
            error_logger.error(f"Customer Insights Generator: SQLite error saving insights to DB for state {current_state.state_id}: {e_db_sqlite}\n{traceback.format_exc()}")
        except Exception as e_db:
            error_logger.error(f"Customer Insights Generator: Failed to save customer insights to database for state {current_state.state_id}: {e_db}\n{traceback.format_exc()}")
        finally:
            if conn:
                conn.close()
        logger.info(f"Customer Insights Generator: Generated {len(parsed_insights)} customer segments for state {current_state.state_id}.")
    except ValueError as ve:
        error_logger.error(f"Customer Insights Generator: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.customer_insights = default_insights
        current_state.node_errors.append({"node": "customer_insights_generator", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e:
        error_logger.error(f"Customer Insights Generator: Failed for state {current_state.state_id}: {e}\n{traceback.format_exc()}")
        current_state.customer_insights = default_insights
        current_state.node_errors.append({"node": "customer_insights_generator", "error": f"Exception: {str(e)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"Customer Insights Generator: Node completed for state {current_state.state_id}.")
    return current_state.model_dump()

async def report_template_generator(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Report Template Generator: Domain='{current_state.market_domain}', UserID='{current_state.user_id or 'N/A'}'")
    default_tmpl = f"# Market Intelligence Report: {current_state.market_domain}\n## Executive Summary\n..."
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.1
        if user_google_api_key:
            logger.info(f"Report Template Generator: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Report Template Generator: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Report Template Generator: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        prompt = ChatPromptTemplate.from_messages([("system", "Create a markdown report template for {market_domain} on query '{query}'. Sections: Title, Date, Prepared By, Executive Summary, Key Trends (name, desc, impact, timeframe), Opportunities (name, desc, potential), Recommendations (title, desc, priority), Competitive Landscape, Visualizations (placeholders like ![Chart Description](filename.png)), Appendix. No \`\`\`markdown\`\`\` fences."),
                                                 ("human", "Generate template for market: {market_domain}, query: {query}")])
        chain = prompt | llm | StrOutputParser()
        generated_template = await chain.ainvoke({"market_domain": current_state.market_domain, "query": current_state.query or "General Overview"})
        cleaned_template = generated_template.strip()
        if cleaned_template.startswith("```markdown"):
            cleaned_template = cleaned_template[len("```markdown"):].strip()
        if cleaned_template.endswith("```"):
            cleaned_template = cleaned_template[:-len("```")].strip()
        current_state.report_template = cleaned_template or default_tmpl
    except ValueError as ve:
        error_logger.error(f"Report Template Generator: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.report_template = default_tmpl
        current_state.node_errors.append({"node": "report_template_generator", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e:
        error_logger.error(f"Report Template Generator: Failed for state {current_state.state_id}: {e}\n{traceback.format_exc()}")
        current_state.report_template = default_tmpl
        current_state.node_errors.append({"node": "report_template_generator", "error": f"Exception: {str(e)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"Report Template Generator: Template length {len(current_state.report_template or '')} for state {current_state.state_id}.")
    return current_state.model_dump()

def get_vector_store_path(current_state: MarketIntelligenceState) -> str:
    base_dir = get_agent_base_reports_dir()
    report_specific_dir = current_state.report_dir or os.path.join(base_dir, f"VS_FALLBACK_{current_state.state_id[:4]}")
    if not os.path.isabs(report_specific_dir):
        report_specific_dir = os.path.join(base_dir, report_specific_dir)
    os.makedirs(report_specific_dir, exist_ok=True)
    return os.path.join(report_specific_dir, f"vector_store_faiss_{current_state.state_id[:4]}")

async def setup_vector_store(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Vector Store Setup: StateID='{current_state.state_id}'")
    if not current_state.report_dir:
        current_state.report_dir = os.path.join(get_agent_base_reports_dir(), f"VS_SETUP_FALLBACK_DIR_{current_state.state_id[:4]}")
        os.makedirs(current_state.report_dir, exist_ok=True)
        logger.warning(f"report_dir was not set, using fallback: {current_state.report_dir}")
    vs_data_json_path = os.path.join(current_state.report_dir, f"{current_state.market_domain.lower().replace(' ', '_')}_data_sources.json")
    docs_for_vs = []
    if os.path.exists(vs_data_json_path):
        try:
            with open(vs_data_json_path, "r", encoding="utf-8") as f:
                data_items = json.load(f)
                for item in data_items:
                    content = item.get('full_content') or item.get('summary', '')
                    if content:
                        docs_for_vs.append({"page_content": content, "metadata": {"source": item.get('source', 'Unknown'), "title": item.get('title', 'Untitled')}})
        except Exception as e_json_read:
            error_logger.error(f"Vector Store Setup: Failed to read JSON '{vs_data_json_path}': {e_json_read}")
    if not docs_for_vs:
        logger.warning("Vector Store Setup: No documents found for vector store. Creating minimal fallback.")
        docs_for_vs = [{"page_content": f"Market Intelligence Report for {current_state.market_domain}. Query: {current_state.query or 'N/A'}.", "metadata": {"source": "Fallback", "title": "Fallback Document"}}]
    try:
        embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
        split_docs = []
        for doc in docs_for_vs:
            chunks = text_splitter.split_text(doc["page_content"])
            for chunk in chunks:
                split_docs.append({"page_content": chunk, "metadata": doc["metadata"]})
        texts = [doc["page_content"] for doc in split_docs]
        metadatas = [doc["metadata"] for doc in split_docs]
        vector_store = FAISS.from_texts(texts, embeddings, metadatas=metadatas)
        vs_path = get_vector_store_path(current_state)
        vector_store.save_local(vs_path)
        current_state.vector_store_path = vs_path
        logger.info(f"Vector Store Setup: Created and saved to '{vs_path}' with {len(split_docs)} chunks.")
    except Exception as e_vs:
        error_logger.error(f"Vector Store Setup: Failed to create vector store: {e_vs}\n{traceback.format_exc()}")
        current_state.vector_store_path = None
        current_state.node_errors.append({"node": "setup_vector_store", "error": f"Exception: {str(e_vs)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info("Vector Store Setup: Node completed.")
    return current_state.model_dump()

async def rag_query_handler(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"RAG Query Handler: StateID='{current_state.state_id}', Question='{current_state.question or 'N/A'}', UserID='{current_state.user_id or 'N/A'}'")
    are_rag_deps_available = True
    try:
        from langchain_community.embeddings import HuggingFaceEmbeddings # Already imported at top
        from langchain_community.vectorstores import FAISS # Already imported at top
        from langchain.chains import RetrievalQA # This was the one causing F401, re-adding for this function
    except ImportError:
        are_rag_deps_available = False
        logger.warning("RAG Query Handler: Dependencies (HuggingFaceEmbeddings, FAISS, RetrievalQA) not available. RAG functionality will be skipped.")
    if not are_rag_deps_available:
        current_state.query_response = "RAG system components are not available due to missing dependencies."
        save_state(current_state)
        return current_state.model_dump()
    if not current_state.question:
        logger.info(f"RAG Query Handler: No question provided for state {current_state.state_id}. Skipping.")
        current_state.query_response = "No question provided for RAG query."
        save_state(current_state)
        return current_state.model_dump()
    if not current_state.vector_store_path or not os.path.exists(current_state.vector_store_path):
        logger.warning(f"RAG Query Handler: Vector store not found at '{current_state.vector_store_path}' for state {current_state.state_id}. Cannot answer question.")
        current_state.query_response = "Vector store not available or not yet created. Cannot answer question."
        save_state(current_state)
        return current_state.model_dump()
    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=current_state.user_id)
        llm_temperature = 0.2
        if user_google_api_key:
            logger.info(f"RAG Query Handler: Using user-provided Google Gemini API key for state {current_state.state_id}")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"RAG Query Handler: Using default Google Gemini API key (from env) for state {current_state.state_id}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"RAG Query Handler: GOOGLE_API_KEY not found in environment for default LLM init for state {current_state.state_id}. LLM calls may fail.")
            llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)
        embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
        vector_store = FAISS.load_local(current_state.vector_store_path, embeddings, allow_dangerous_deserialization=True)
        qa_chain = RetrievalQA.from_chain_type(llm=llm, chain_type="stuff", retriever=vector_store.as_retriever(search_kwargs={"k": 5}))
        response = await qa_chain.ainvoke({"query": current_state.question})
        current_state.query_response = response.get("result", "No response generated.")
        logger.info(f"RAG Query Handler: Generated response for question: '{current_state.question}' for state {current_state.state_id}")
    except ValueError as ve:
        error_logger.error(f"RAG Query Handler: Value error for state {current_state.state_id} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        current_state.query_response = f"Error processing question due to configuration: {str(ve)}"
        current_state.node_errors.append({"node": "rag_query_handler", "error": f"ValueError: {str(ve)}", "details": traceback.format_exc(limit=2)})
    except Exception as e_rag:
        error_logger.error(f"RAG Query Handler: Failed to process question '{current_state.question}' for state {current_state.state_id}: {e_rag}\n{traceback.format_exc()}")
        current_state.query_response = f"Error processing question: {str(e_rag)}"
        current_state.node_errors.append({"node": "rag_query_handler", "error": f"Exception: {str(e_rag)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info(f"RAG Query Handler: Node completed for state {current_state.state_id}.")
    return current_state.model_dump()

def generate_charts(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Chart Generator: StateID='{current_state.state_id}'")
    if plt is None or sns is None or pd is None or np is None:
        logger.warning("Chart Generator: Required libraries not available. Skipping chart generation.")
        return current_state.model_dump()
    if not current_state.report_dir:
        logger.warning("Chart Generator: No report directory set. Cannot save charts.")
        return current_state.model_dump()
    try:
        sns.set_style("whitegrid")
        plt.rcParams['figure.figsize'] = (10, 6)
        plt.rcParams['font.size'] = 10
        if current_state.market_trends:
            trend_names = [t.get('trend_name', 'Unknown') for t in current_state.market_trends[:5]]
            impact_values = []
            for t in current_state.market_trends[:5]:
                impact = t.get('estimated_impact', 'Medium')
                if impact == 'High':
                    impact_values.append(3)
                elif impact == 'Medium':
                    impact_values.append(2)
                else:
                    impact_values.append(1)
            plt.figure(figsize=(12, 6))
            bars = plt.bar(trend_names, impact_values, color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'])
            plt.title(f'Market Trends Impact Analysis - {current_state.market_domain}', fontsize=14, fontweight='bold')
            plt.xlabel('Trends', fontsize=12)
            plt.ylabel('Impact Level', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.yticks([1, 2, 3], ['Low', 'Medium', 'High'])
            plt.tight_layout()
            for bar, value in zip(bars, impact_values):
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height + 0.05, ['Low', 'Medium', 'High'][int(value)-1], ha='center', va='bottom', fontweight='bold')
            chart1_path = os.path.join(current_state.report_dir, "market_trends_impact.png")
            plt.savefig(chart1_path, dpi=300, bbox_inches='tight')
            plt.close()
            current_state.chart_paths.append(chart1_path)
            logger.info(f"Chart Generator: Saved trends chart to {chart1_path}")
        if current_state.opportunities:
            opp_names = [o.get('opportunity_name', 'Unknown') for o in current_state.opportunities[:5]]
            potential_values = []
            for o in current_state.opportunities[:5]:
                potential = o.get('estimated_potential', 'Medium')
                if potential == 'High' or potential == 'Very High':
                    potential_values.append(3)
                elif potential == 'Medium':
                    potential_values.append(2)
                else:
                    potential_values.append(1)
            plt.figure(figsize=(10, 8))
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
            wedges, texts, autotexts = plt.pie(potential_values, labels=opp_names, autopct='%1.1f%%', colors=colors, startangle=90)
            plt.title(f'Market Opportunities Distribution - {current_state.market_domain}', fontsize=14, fontweight='bold')
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
            chart2_path = os.path.join(current_state.report_dir, "opportunities_distribution.png")
            plt.savefig(chart2_path, dpi=300, bbox_inches='tight')
            plt.close()
            current_state.chart_paths.append(chart2_path)
            logger.info(f"Chart Generator: Saved opportunities chart to {chart2_path}")
        if current_state.customer_insights:
            segment_names = [c.get('segment_name', 'Unknown') for c in current_state.customer_insights[:5]]
            percentages = [c.get('percentage', 0) for c in current_state.customer_insights[:5]]
            satisfaction_scores = [c.get('satisfaction_score', 0) for c in current_state.customer_insights[:5]]
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
            ax1.pie(percentages, labels=segment_names, autopct='%1.1f%%', colors=colors, startangle=90)
            ax1.set_title('Customer Segment Distribution', fontsize=12, fontweight='bold')
            bars = ax2.bar(segment_names, satisfaction_scores, color=colors)
            ax2.set_title('Customer Satisfaction by Segment', fontsize=12, fontweight='bold')
            ax2.set_ylabel('Satisfaction Score (1-10)')
            ax2.set_ylim(0, 10)
            for bar, score in zip(bars, satisfaction_scores):
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + 0.1, f'{score:.1f}', ha='center', va='bottom', fontweight='bold')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            chart3_path = os.path.join(current_state.report_dir, "customer_insights.png")
            plt.savefig(chart3_path, dpi=300, bbox_inches='tight')
            plt.close()
            current_state.chart_paths.append(chart3_path)
            logger.info(f"Chart Generator: Saved customer insights chart to {chart3_path}")
        if current_state.strategic_recommendations:
            strategy_names = [s.get('strategy_title', 'Unknown') for s in current_state.strategic_recommendations[:5]]
            priority_values = []
            for s in current_state.strategic_recommendations[:5]:
                priority = s.get('priority_level', 'Medium')
                if priority == 'High' or priority == 'Critical':
                    priority_values.append(3)
                elif priority == 'Medium':
                    priority_values.append(2)
                else:
                    priority_values.append(1)
            plt.figure(figsize=(12, 6))
            colors = ['#FF6B6B' if p == 3 else '#FFEAA7' if p == 2 else '#96CEB4' for p in priority_values]
            bars = plt.barh(strategy_names, priority_values, color=colors)
            plt.title(f'Strategic Recommendations Priority - {current_state.market_domain}', fontsize=14, fontweight='bold')
            plt.xlabel('Priority Level', fontsize=12)
            plt.xticks([1, 2, 3], ['Low', 'Medium', 'High'])
            for bar, value in zip(bars, priority_values):
                width = bar.get_width()
                plt.text(width + 0.05, bar.get_y() + bar.get_height()/2., ['Low', 'Medium', 'High'][int(value)-1], ha='left', va='center', fontweight='bold')
            plt.tight_layout()
            chart4_path = os.path.join(current_state.report_dir, "strategic_recommendations.png")
            plt.savefig(chart4_path, dpi=300, bbox_inches='tight')
            plt.close()
            current_state.chart_paths.append(chart4_path)
            logger.info(f"Chart Generator: Saved strategies chart to {chart4_path}")
        current_state.download_files["charts"] = current_state.chart_paths
        logger.info(f"Chart Generator: Generated {len(current_state.chart_paths)} charts.")
    except Exception as e_chart:
        error_logger.error(f"Chart Generator: Failed to generate charts: {e_chart}\n{traceback.format_exc()}")
        current_state.node_errors.append({"node": "generate_charts", "error": f"Exception: {str(e_chart)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    return current_state.model_dump()

async def final_report_generator(current_state: MarketIntelligenceState) -> Dict[str, Any]:
    logger.info(f"Final Report Generator: StateID='{current_state.state_id}'")
    if not current_state.report_dir:
        logger.warning("Final Report Generator: No report directory set. Cannot save report.")
        return current_state.model_dump()
    try:
        report_content = current_state.report_template or f"# Market Intelligence Report: {current_state.market_domain}\n\nNo template available."
        
        # Replace placeholders with actual data
        current_date = datetime.now().strftime("%B %d, %Y")
        report_content = report_content.replace("{{DATE}}", current_date)
        report_content = report_content.replace("{{MARKET_DOMAIN}}", current_state.market_domain)
        report_content = report_content.replace("{{QUERY}}", current_state.query or "General Analysis")
        
        # Add trends section
        if current_state.market_trends:
            trends_section = "\n## Key Market Trends\n\n"
            for i, trend in enumerate(current_state.market_trends[:5], 1):
                trends_section += f"### {i}. {trend.get('trend_name', 'Unknown Trend')}\n"
                trends_section += f"**Description:** {trend.get('description', 'N/A')}\n\n"
                trends_section += f"**Impact:** {trend.get('estimated_impact', 'Unknown')}\n\n"
                trends_section += f"**Timeframe:** {trend.get('timeframe', 'Unknown')}\n\n"
                if trend.get('supporting_evidence'):
                    trends_section += f"**Evidence:** {trend.get('supporting_evidence')}\n\n"
                trends_section += "---\n\n"
            report_content += trends_section
        if current_state.opportunities:
            opportunities_section = "\n## Market Opportunities\n\n"
            for i, opp in enumerate(current_state.opportunities[:5], 1):
                opportunities_section += f"### {i}. {opp.get('opportunity_name', 'Unknown Opportunity')}\n"
                opportunities_section += f"**Description:** {opp.get('description', 'N/A')}\n\n"
                opportunities_section += f"**Potential:** {opp.get('estimated_potential', 'Unknown')}\n\n"
                if opp.get('target_segment'):
                    opportunities_section += f"**Target Segment:** {opp.get('target_segment')}\n\n"
                opportunities_section += "---\n\n"
            report_content += opportunities_section
        if current_state.customer_insights:
            insights_section = "\n## Customer Insights\n\n"
            for i, insight in enumerate(current_state.customer_insights[:5], 1):
                insights_section += f"### {i}. {insight.get('segment_name', 'Unknown Segment')}\n"
                insights_section += f"**Description:** {insight.get('description', 'N/A')}\n\n"
                insights_section += f"**Market Share:** {insight.get('percentage', 0)}%\n\n"
                insights_section += f"**Satisfaction Score:** {insight.get('satisfaction_score', 'N/A')}/10\n\n"
                insights_section += f"**Growth Potential:** {insight.get('growth_potential', 'Unknown')}\n\n"
                if insight.get('pain_points'):
                    insights_section += f"**Key Pain Points:**\n"
                    for pain_point in insight.get('pain_points', []):
                        insights_section += f"- {pain_point}\n"
                    insights_section += "\n"
                insights_section += "---\n\n"
            report_content += insights_section
        if current_state.strategic_recommendations:
            strategies_section = "\n## Strategic Recommendations\n\n"
            for i, strategy in enumerate(current_state.strategic_recommendations[:5], 1):
                strategies_section += f"### {i}. {strategy.get('strategy_title', 'Unknown Strategy')}\n"
                strategies_section += f"**Description:** {strategy.get('description', 'N/A')}\n\n"
                strategies_section += f"**Priority:** {strategy.get('priority_level', 'Unknown')}\n\n"
                if strategy.get('expected_outcome'):
                    strategies_section += f"**Expected Outcome:** {strategy.get('expected_outcome')}\n\n"
                if strategy.get('implementation_steps'):
                    strategies_section += f"**Implementation Steps:**\n"
                    for step in strategy.get('implementation_steps', []):
                        strategies_section += f"- {step}\n"
                    strategies_section += "\n"
                strategies_section += "---\n\n"
            report_content += strategies_section
        if current_state.chart_paths:
            charts_section = "\n## Visualizations\n\n"
            for chart_path in current_state.chart_paths:
                chart_name = os.path.basename(chart_path).replace('.png', '').replace('_', ' ').title()
                charts_section += f"### {chart_name}\n"
                charts_section += f"![{chart_name}]({os.path.basename(chart_path)})\n\n"
            report_content += charts_section
        if current_state.query_response and current_state.question:
            rag_section = f"\n## Analysis Response\n\n"
            rag_section += f"**Question:** {current_state.question}\n\n"
            rag_section += f"**Answer:** {current_state.query_response}\n\n"
            report_content += rag_section
        report_path = os.path.join(current_state.report_dir, "market_intelligence_report.md")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_content)
        current_state.download_files["final_report"] = report_path
        logger.info(f"Final Report Generator: Saved report to {report_path}")
        readme_content = f"# Market Intelligence Report - {current_state.market_domain}\n\nGenerated on: {current_date}\nQuery: {current_state.query or 'General Analysis'}\n\n## Files in this Report\n\n- `market_intelligence_report.md` - Complete market intelligence report\n- `{current_state.market_domain.lower().replace(' ', '_')}_data_sources.json` - Raw data sources\n- `{current_state.market_domain.lower().replace(' ', '_')}_data_sources.csv` - Raw data in CSV format\n- `market_trends.json` - Identified market trends\n- `opportunities.json` - Market opportunities\n- `customer_insights.json` - Customer segment analysis\n- `strategies.json` - Strategic recommendations\n\n## Charts Generated\n\n"
        for chart_path in current_state.chart_paths:
            chart_name = os.path.basename(chart_path).replace('.png', '').replace('_', ' ').title()
            readme_content += f"- `{os.path.basename(chart_path)}` - {chart_name}\n"
        readme_content += f"\n## Summary\n\n- **Trends Identified:** {len(current_state.market_trends)}\n- **Opportunities Found:** {len(current_state.opportunities)}\n- **Customer Segments:** {len(current_state.customer_insights)}\n- **Strategic Recommendations:** {len(current_state.strategic_recommendations)}\n- **Charts Generated:** {len(current_state.chart_paths)}\n\nFor detailed analysis, see the complete report in `market_intelligence_report.md`.\n"
        readme_path = os.path.join(current_state.report_dir, "README.md")
        with open(readme_path, "w", encoding="utf-8") as f:
            f.write(readme_content)
        current_state.download_files["readme"] = readme_path
        logger.info(f"Final Report Generator: Saved README to {readme_path}")

    except Exception as e_report:
        error_logger.error(f"Final Report Generator: Failed to generate report: {e_report}\n{traceback.format_exc()}")
        current_state.node_errors.append({"node": "final_report_generator", "error": f"Exception: {str(e_report)}", "details": traceback.format_exc(limit=2)})
    save_state(current_state)
    logger.info("Final Report Generator: Node completed.")
    return current_state.model_dump()

# MarketIntelligenceAgent class for compatibility with main.py
class MarketIntelligenceAgent:
    def __init__(self):
        self.state = None
        logger.info("MarketIntelligenceAgent initialized")

    async def run_analysis(self, query: str, market_domain: str, question: str = None) -> Dict[str, Any]:
        return await run_market_intelligence_agent(query, market_domain, question)

    async def chat(
        self, message: str, session_id: str, history: List[Dict[str, Any]] = None, user_id: Optional[str] = None
    ) -> str:
        if history is None:
            history = load_chat_history(session_id)
        return await chat_with_agent(message, session_id, history, user_id=user_id)

    def get_state(self, state_id: str) -> Optional[MarketIntelligenceState]:
        return load_state(state_id)

    def get_customer_insights(self, state_id: str) -> List[Dict[str, Any]]:
        state = load_state(state_id)
        if state:
            return state.customer_insights
        return []

    def prepare_download(self, state_id: str, category: str) -> Optional[str]:
        state = load_state(state_id)
        if not state or not state.download_files:
            return None
        db_path = get_db_path()
        try:
            conn = sqlite3.connect(db_path)
            cursor_obj = conn.cursor()
            download_id = str(uuid4())
            file_path = state.download_files.get(category)
            if file_path:
                cursor_obj.execute(
                    "INSERT INTO downloads (id, state_id, category, file_path, file_type) VALUES (?, ?, ?, ?, ?)",
                    (download_id, state_id, category, file_path, os.path.splitext(file_path)[1]),
                )
                conn.commit()
                conn.close()
                return file_path
        except Exception as e:
            error_logger.error(f"Failed to prepare download: {e}")
        return None


async def run_market_intelligence_agent(
    query_str: str, market_domain_str: str, question_str: Optional[str] = None, user_id: Optional[str] = None, uploaded_document_ids: Optional[List[str]] = None
) -> Dict[str, Any]:
    logger.info(
        f"Agent Run: Starting with Query='{query_str}', Domain='{market_domain_str}', Question='{question_str or 'N/A'}', UserID='{user_id or 'N/A'}', UploadedDocIDs='{uploaded_document_ids or 'None'}'"
    )
    error_state_id = str(uuid4())
    error_report_dir_path = None
    error_report_file_path = None
    try:
        initial_state = MarketIntelligenceState(
            market_domain=market_domain_str,
            query=query_str,
            question=question_str,
            user_id=user_id,
            uploaded_document_ids=uploaded_document_ids
        )
        logger.info(f"Agent Run: Initial state created with ID: {initial_state.state_id} for UserID: {user_id}, includes {len(uploaded_document_ids or [])} uploaded doc IDs.")
        workflow = StateGraph(dict)
        workflow.add_node("market_data_collector", market_data_collector)
        workflow.add_node("trend_analyzer", trend_analyzer)
        workflow.add_node("opportunity_identifier", opportunity_identifier)
        workflow.add_node("strategy_recommender", strategy_recommender)
        workflow.add_node("customer_insights_generator", customer_insights_generator)
        workflow.add_node("report_template_generator", report_template_generator)
        workflow.add_node("setup_vector_store", setup_vector_store)
        workflow.add_node("rag_query_handler", rag_query_handler)
        workflow.add_node("generate_charts", generate_charts)
        workflow.add_node("final_report_generator", final_report_generator)
        workflow.set_entry_point("market_data_collector")
        workflow.add_edge("market_data_collector", "trend_analyzer")
        workflow.add_edge("trend_analyzer", "opportunity_identifier")
        workflow.add_edge("opportunity_identifier", "strategy_recommender")
        workflow.add_edge("strategy_recommender", "customer_insights_generator")
        workflow.add_edge("customer_insights_generator", "report_template_generator")
        workflow.add_edge("report_template_generator", "setup_vector_store")
        workflow.add_edge("setup_vector_store", "rag_query_handler")
        workflow.add_edge("rag_query_handler", "generate_charts")
        workflow.add_edge("generate_charts", "final_report_generator")
        workflow.add_edge("final_report_generator", END)
        app = workflow.compile()
        final_state_dict = await app.ainvoke(initial_state.model_dump())
        final_state = MarketIntelligenceState(**final_state_dict)
        report_dir_relative = None
        if final_state.report_dir:
            if final_state.report_dir.startswith("/tmp/"):
                report_dir_relative = os.path.relpath(final_state.report_dir, "/tmp")
            else:
                base_reports_dir = get_agent_base_reports_dir()
                abs_report_dir = os.path.abspath(final_state.report_dir)
                if abs_report_dir.startswith(os.path.abspath(base_reports_dir)):
                    report_dir_relative = os.path.relpath(abs_report_dir, base_reports_dir)
                else:
                    report_dir_relative = os.path.basename(abs_report_dir)
        chart_filenames = [os.path.basename(chart_path) for chart_path in final_state.chart_paths]
        return_data = {
            "success": True,
            "state_id": final_state.state_id,
            "report_dir_relative": report_dir_relative,
            "report_filename": "market_intelligence_report.md",
            "chart_filenames": chart_filenames,
            "data_json_filename": f"{final_state.market_domain.lower().replace(' ', '_')}_data_sources.json",
            "data_csv_filename": f"{final_state.market_domain.lower().replace(' ', '_')}_data_sources.csv",
            "readme_filename": "README.md",
            "log_filename": "market_intelligence.log",
            "rag_log_filename": "market_intelligence_errors.log",
            "vector_store_dirname": os.path.basename(final_state.vector_store_path) if final_state.vector_store_path else None,
            "query_response": final_state.query_response,
            "download_files": final_state.download_files,
            "node_errors": final_state.node_errors # Include node-specific errors
        }

        logger.info(f"Agent Run: Successfully completed for StateID '{final_state.state_id}'. Node errors encountered: {len(final_state.node_errors)}")
        return return_data

    except Exception as e_agent_run:
        error_logger.critical(f"Agent Run: CRITICAL FAILURE: {e_agent_run}\n{traceback.format_exc()}")
        tb_str = traceback.format_exc()
        try:
            error_report_dir_path = os.path.join(get_agent_base_reports_dir(), f"ERROR_REPORT_{error_state_id[:8]}")
            os.makedirs(error_report_dir_path, exist_ok=True)
            error_report_file_path = os.path.join(error_report_dir_path, f"ERROR_REPORT_{error_state_id[:8]}.md")
            with open(error_report_file_path, "w", encoding="utf-8") as f:
                f.write(f"""# Market Intelligence Agent - Error Report

**Error ID:** {error_state_id}
**Timestamp:** {datetime.now().isoformat()}
**Query:** {query_str}
**Market Domain:** {market_domain_str}
**Question:** {question_str or 'N/A'}

## Error Details

""")
                f.write(f"\`\`\`\n{tb_str}\n\`\`\`")
        except Exception as e_error_report:
            error_logger.error(f"Agent Run: Failed to write error report: {e_error_report}")

        return {
            "success": False,
            "state_id": error_state_id,
            "report_dir_relative": os.path.relpath(error_report_dir_path, get_agent_base_reports_dir()) if error_report_dir_path else None,
            "report_filename": os.path.basename(error_report_file_path) if error_report_file_path else None,
            "chart_filenames": [],
            "data_json_filename": None,
            "data_csv_filename": None,
            "readme_filename": None,
            "log_filename": None,
            "rag_log_filename": None,
            "vector_store_dirname": None,
            "query_response": None,
            "download_files": None,
            "error": str(e_agent_run)
        }


async def chat_with_agent(
    message: str,
    session_id: str,
    history: List[Dict[str, Any]],
    user_id: Optional[str] = None,
    request_context: Optional[Dict[str, Any]] = None # For file context
) -> str:
    logger.info(f"Agent Chat: SessionID {session_id}, UserID: {user_id or 'N/A'}, Message: '{message[:100]}...', Context: {request_context is not None}")
    save_chat_message(session_id, "user", message)

    langchain_history = []
    for msg_data in history:
        if msg_data["type"] == "user":
            langchain_history.append(HumanMessage(content=msg_data["content"]))
        elif msg_data["type"] == "ai":
            langchain_history.append(AIMessage(content=msg_data["content"]))

    context_str_for_llm = ""
    context_files_summary = "No file context provided."

    if request_context and request_context.get("files") and user_id:
        files_in_context = request_context.get("files", [])
        if files_in_context:
            logger.info(f"Chat RAG: Processing {len(files_in_context)} files for context for user {user_id}.")
            context_files_summary = f"Using context from {len(files_in_context)} file(s): {', '.join([f.get('filename', f.get('file_id', 'unknown')) for f in files_in_context])}"

            all_file_texts = []
            supabase = get_supabase_client()
            if supabase:
                for file_info in files_in_context:
                    file_id = file_info.get("file_id")
                    if file_id:
                        try:
                            # Ensure user owns the document they are trying to use in context
                            doc_response = supabase.table("documents").select("text, original_filename").eq("id", file_id).eq("uploader_id", user_id).maybe_single().execute()
                            if doc_response.data and doc_response.data.get("text"):
                                all_file_texts.append(doc_response.data["text"])
                                logger.debug(f"Chat RAG: Loaded text for file_id {file_id} (filename: {doc_response.data.get('original_filename', 'N/A')})")
                            else:
                                logger.warning(f"Chat RAG: File {file_id} not found or no text content for user {user_id}.")
                        except Exception as e_fetch_doc_text:
                            error_logger.error(f"Chat RAG: Error fetching text for file {file_id}: {e_fetch_doc_text}")

            if all_file_texts:
                try:
                    # Combine all texts and create an on-the-fly vector store
                    combined_text = "\n\n--- (New Document Context) ---\n\n".join(all_file_texts)
                    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1500, chunk_overlap=200)
                    chunks = text_splitter.split_text(combined_text)

                    if chunks:
                        logger.info(f"Chat RAG: Created {len(chunks)} chunks from {len(all_file_texts)} files for in-memory FAISS.")
                        embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
                        # Allow dangerous deserialization for FAISS if it's loaded from disk, not strictly needed for from_texts
                        vector_store = FAISS.from_texts(texts=chunks, embedding=embeddings)

                        # Retrieve top N relevant chunks based on the user's message
                        retriever = vector_store.as_retriever(search_kwargs={"k": 3}) # Get top 3 chunks
                        relevant_docs = retriever.get_relevant_documents(message)

                        if relevant_docs:
                            context_str_for_llm = "\n\n--- Relevant Context from Uploaded Files ---\n"
                            for i, doc_chunk in enumerate(relevant_docs):
                                context_str_for_llm += f"\n[Excerpt {i+1} from files]:\n{doc_chunk.page_content}\n---"
                            logger.info(f"Chat RAG: Retrieved {len(relevant_docs)} relevant chunks for the query.")
                        else:
                            logger.info("Chat RAG: No relevant chunks found in provided files for the query.")
                            context_str_for_llm = "\n\n--- Context from Uploaded Files (No specific relevant excerpts found, general content considered) ---"
                            # Optionally, include a summary or just the fact that files were considered.
                    else:
                        logger.warning("Chat RAG: No text chunks generated from files, skipping vector store creation.")
                except Exception as e_rag_process:
                    error_logger.error(f"Chat RAG: Error during in-memory RAG processing: {e_rag_process}\n{traceback.format_exc()}")
                    context_str_for_llm = "\n\n[Error processing file context for RAG]\n"

    try:
        user_google_api_key = get_api_key("GOOGLE_GEMINI", user_id=user_id)
        llm_temperature = 0.7

        if user_google_api_key:
            logger.info(f"Chat: Using user-provided Google Gemini API key for session {session_id}, UserID {user_id}")
            chat_llm = ChatGoogleGenerativeAI(model="gemini-pro", google_api_key=user_google_api_key, temperature=llm_temperature)
        else:
            logger.info(f"Chat: Using default Google Gemini API key (from env) for session {session_id}, UserID {user_id or 'N/A (fallback)'}")
            if not os.getenv("GOOGLE_API_KEY"):
                error_logger.warning(f"Chat: GOOGLE_API_KEY not found in environment for default LLM init for session {session_id}.")
            chat_llm = ChatGoogleGenerativeAI(model="gemini-pro", temperature=llm_temperature)

        system_message_content = "You are a helpful assistant. Respond to the user's query. "
        if context_str_for_llm:
            system_message_content += f"Base your answer primarily on the following context from uploaded files if relevant, otherwise indicate that the files were not relevant to the query. \n{context_files_summary}\n{context_str_for_llm}"
        else:
            system_message_content += "Use your general knowledge if no specific file context is provided or relevant."

        prompt_template = ChatPromptTemplate.from_messages([
            ("system", system_message_content),
            MessagesPlaceholder(variable_name="chat_history"),
            ("human", "{input}")
        ])
        chain = prompt_template | chat_llm | StrOutputParser()

        response_text = await chain.ainvoke({"input": message, "chat_history": langchain_history})

        save_chat_message(session_id, "ai", response_text)
        logger.info(f"Agent Chat: Response generated for session_id {session_id}, UserID {user_id or 'N/A'}. File context used: {bool(context_str_for_llm)}")
        return response_text

    except ValueError as ve:
        error_logger.error(f"Agent Chat: Value error for session {session_id}, UserID {user_id or 'N/A'} (possibly API key issue): {ve}\n{traceback.format_exc()}")
        error_response = "Sorry, I encountered a configuration error while processing your message."
        save_chat_message(session_id, "ai", error_response)
        return error_response
    except Exception as e:
        error_logger.error(f"Agent Chat: Error processing message for session {session_id}, UserID {user_id or 'N/A'}: {e}\n{traceback.format_exc()}")
        error_response = "Sorry, I encountered an unexpected error while processing your message."
        save_chat_message(session_id, "ai", error_response)
        return error_response

if __name__ == "__main__":
    cmd_arg_parser = argparse.ArgumentParser(description="Market Intelligence Agent CLI")
    cmd_arg_parser.add_argument("--query", type=str, default="AI impact on EdTech", help="The main query or topic for market analysis.")
    cmd_arg_parser.add_argument("--market", type=str, default="EdTech", help="The target market domain for analysis.")
    cmd_arg_parser.add_argument("--question", type=str, default=None, help="Optional: A specific question for the RAG system about the generated data/report.")

    parsed_cli_args = cmd_arg_parser.parse_args()

    logger.info(f"Agent CLI: Starting with Query='{parsed_cli_args.query}', Market='{parsed_cli_args.market}', Question='{parsed_cli_args.question or 'N/A'}'")
    # Note: This local CLI runner will need to be adapted to run an async function,
    # e.g., using asyncio.run()
    import asyncio
    cli_run_output = asyncio.run(run_market_intelligence_agent(
        query_str=parsed_cli_args.query,
        market_domain_str=parsed_cli_args.market,
        question_str=parsed_cli_args.question
    ))

    print("--- Agent CLI Run Summary ---")
def get_latest_user_state_id(user_id: str) -> Optional[str]:
    """
    Retrieves the ID of the most recent analysis state for a given user
    from the agent's local SQLite database.
    Assumes 'created_at' column exists and stores ISO8601 timestamps for correct sorting.
    """
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        # Ensure created_at is being stored in a sortable format (ISO8601 string is fine with DATETIME())
        cursor.execute(
            "SELECT id FROM states WHERE user_id = ? ORDER BY DATETIME(created_at) DESC LIMIT 1",
            (user_id,)
        )
        row = cursor.fetchone()
        if row:
            logger.info(f"Found latest state ID '{row[0]}' for user_id '{user_id}'.")
            return row[0]
        else:
            logger.info(f"No states found for user_id '{user_id}'.")
            return None
    except sqlite3.Error as e:
        error_logger.error(f"SQLite error fetching latest state ID for user {user_id}: {e}")
        return None
    except Exception as e: # Catch any other unexpected errors
        error_logger.error(f"Unexpected error fetching latest state ID for user {user_id}: {e}\n{traceback.format_exc()}")
        return None
    finally:
        if conn:
            conn.close()
    print(f"Success: {cli_run_output.get('success')}")
    print(f"State ID: {cli_run_output.get('state_id')}")
    print(f"Report Directory (relative to /tmp or api_python/reports1): {cli_run_output.get('report_dir_relative')}")
    print(f"Report Filename: {cli_run_output.get('report_filename')}")
    if cli_run_output.get("query_response"):
        print(f"RAG Query Response: {cli_run_output.get('query_response')}")
    if cli_run_output.get("error"):
        print(f"Error Message: {cli_run_output.get('error')}")

    if cli_run_output.get("success"):
        print("To view results, check the 'reports1' directory (likely in '/tmp/reports1/' on Vercel or 'api_python/reports1/' locally), then find the subdirectory indicated by 'report_dir_relative'.")
    else:
        print("Agent run encountered errors. Please check logs ('market_intelligence.log', 'market_intelligence_errors.log') in the Python function's log (Vercel) or 'api_python/' directory (local) for details.")
